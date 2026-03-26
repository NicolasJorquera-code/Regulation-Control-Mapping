"""Excel register analyzer — extracts a structured summary from any control register.

Parses uploaded Excel files with flexible header matching, extracting
unique control types, business units, sections, placements, methods,
frequencies, roles, systems, regulatory mentions, and sample descriptions.
The resulting ``RegisterSummary`` feeds the ``ConfigProposerAgent``.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)


# ── Header synonym mapping ────────────────────────────────────────────────────

# Maps canonical field names to known header variations (all lowercase).
HEADER_SYNONYMS: dict[str, list[str]] = {
    "control_type": [
        "control_type",
        "control type",
        "selected_level_2",
        "type",
        "type of control",
        "control category",
        "control type name",
        "controltype",
        "ctrl_type",
    ],
    "placement": [
        "selected_level_1",
        "placement",
        "category",
        "control placement",
        "level_1",
        "level 1",
        "preventive or detective",
    ],
    "method": [
        "method",
        "control method",
        "execution method",
        "control_method",
        "automated/manual",
    ],
    "business_unit_id": [
        "business_unit_id",
        "bu_id",
        "business unit id",
        "unit_id",
    ],
    "business_unit_name": [
        "business_unit_name",
        "business unit",
        "bu_name",
        "department",
        "division",
        "unit",
        "business unit name",
        "bu name",
    ],
    "section_id": [
        "hierarchy_id",
        "section_id",
        "section id",
        "section",
        "process_area_id",
        "process area",
    ],
    "section_name": [
        "leaf_name",
        "section_name",
        "section name",
        "process_area_name",
        "process area name",
    ],
    "frequency": [
        "frequency",
        "when",
        "timing",
        "how often",
        "control frequency",
        "schedule",
    ],
    "who": [
        "who",
        "responsible",
        "responsible party",
        "owner",
        "performer",
        "control_owner",
        "control owner",
    ],
    "where": [
        "where",
        "system",
        "application",
        "platform",
        "location",
    ],
    "why": [
        "why",
        "purpose",
        "objective",
        "risk",
        "rationale",
    ],
    "full_description": [
        "full_description",
        "full description",
        "description",
        "control description",
        "narrative",
        "control narrative",
        "control_description",
    ],
    "evidence": [
        "evidence",
        "evidence_artifact",
        "evidence artifact",
        "evidence required",
        "documentation",
    ],
    "control_id": [
        "control_id",
        "control id",
        "id",
        "ctrl_id",
        "ctrl id",
    ],
    "quality_rating": [
        "quality_rating",
        "quality rating",
        "rating",
        "effectiveness",
        "quality",
        "risk level",
    ],
    "regulatory_framework": [
        "regulatory_framework",
        "regulatory framework",
        "regulation",
        "framework",
        "compliance",
    ],
}


# ── Summary model ─────────────────────────────────────────────────────────────


class RegisterSummary(BaseModel):
    """Structured summary of an uploaded control register."""

    row_count: int = 0
    sheet_names: list[str] = Field(default_factory=list)
    column_names: list[str] = Field(default_factory=list)
    raw_headers: list[str] = Field(default_factory=list)
    header_mapping: dict[str, str] = Field(default_factory=dict)

    unique_control_types: list[str] = Field(default_factory=list)
    unique_business_units: list[dict[str, str]] = Field(default_factory=list)
    unique_sections: list[dict[str, str]] = Field(default_factory=list)
    unique_placements: list[str] = Field(default_factory=list)
    unique_methods: list[str] = Field(default_factory=list)
    frequency_values: list[str] = Field(default_factory=list)
    role_mentions: list[str] = Field(default_factory=list)
    system_mentions: list[str] = Field(default_factory=list)
    regulatory_mentions: list[str] = Field(default_factory=list)
    sample_descriptions: list[str] = Field(default_factory=list)


# ── Header matching ───────────────────────────────────────────────────────────


def _match_headers(raw_headers: list[str]) -> dict[str, int]:
    """Match raw Excel headers to canonical field names.

    Returns a mapping of ``{canonical_name: column_index}``.
    """
    mapping: dict[str, int] = {}
    normalised = [h.strip().lower() for h in raw_headers]

    for canonical, synonyms in HEADER_SYNONYMS.items():
        for syn in synonyms:
            if syn in normalised:
                idx = normalised.index(syn)
                if canonical not in mapping:
                    mapping[canonical] = idx
                break
    return mapping


# ── Main analyzer ─────────────────────────────────────────────────────────────


def analyze_register(path: Path | str) -> RegisterSummary:
    """Parse an Excel register and extract a structured summary.

    Reads all sheets (not limited to ``section_*``), detects column
    roles via fuzzy header matching, and extracts unique values across
    all sheets.

    Args:
        path: Path to the ``.xlsx`` file.

    Returns:
        A ``RegisterSummary`` with aggregated data from all sheets.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    all_raw_headers: list[str] = []
    all_canonical: set[str] = set()
    total_rows = 0

    control_types: set[str] = set()
    bus: dict[str, str] = {}  # id -> name  or  name -> name
    sections: dict[str, str] = {}  # id -> name
    placements: set[str] = set()
    methods: set[str] = set()
    frequencies: set[str] = set()
    roles: set[str] = set()
    systems: set[str] = set()
    regulatory: set[str] = set()
    descriptions: list[str] = []
    header_mapping_display: dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        raw_header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        if not all_raw_headers:
            all_raw_headers = raw_header_row

        col_map = _match_headers(raw_header_row)
        if not col_map:
            continue

        for canonical, idx in col_map.items():
            all_canonical.add(canonical)
            if canonical not in header_mapping_display:
                header_mapping_display[canonical] = raw_header_row[idx]

        for row in rows[1:]:
            total_rows += 1

            def _get(field: str) -> str:
                idx = col_map.get(field)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
                return ""

            ct = _get("control_type")
            if ct:
                control_types.add(ct)

            bu_id = _get("business_unit_id")
            bu_name = _get("business_unit_name")
            if bu_name:
                key = bu_id or bu_name
                bus[key] = bu_name
            elif bu_id:
                bus[bu_id] = bu_id

            sec_id = _get("section_id")
            sec_name = _get("section_name")
            if sec_id:
                # Extract top-level section (e.g. "1.0" from "1.0.2.3")
                top = sec_id.split(".")[0]
                top_id = f"{top}.0" if "." not in sec_id or sec_id != f"{top}.0" else sec_id
                if top_id not in sections:
                    sections[top_id] = sec_name or top_id

            placement = _get("placement")
            if placement:
                placements.add(placement)

            method = _get("method")
            if method:
                methods.add(method)

            freq = _get("frequency")
            if freq:
                frequencies.add(freq)

            who = _get("who")
            if who:
                roles.add(who)

            where = _get("where")
            if where:
                systems.add(where)

            reg = _get("regulatory_framework")
            if reg:
                regulatory.add(reg)

            desc = _get("full_description")
            if desc and len(descriptions) < 10:
                descriptions.append(desc)

    wb.close()

    unique_bus = [{"id": k, "name": v} for k, v in sorted(bus.items())]
    unique_sections = [{"id": k, "name": v} for k, v in sorted(sections.items(), key=lambda x: x[0])]

    return RegisterSummary(
        row_count=total_rows,
        sheet_names=wb.sheetnames,
        column_names=list(all_canonical),
        raw_headers=all_raw_headers,
        header_mapping=header_mapping_display,
        unique_control_types=sorted(control_types),
        unique_business_units=unique_bus,
        unique_sections=unique_sections,
        unique_placements=sorted(placements),
        unique_methods=sorted(methods),
        frequency_values=sorted(frequencies),
        role_mentions=sorted(roles),
        system_mentions=sorted(systems),
        regulatory_mentions=sorted(regulatory),
        sample_descriptions=descriptions,
    )
