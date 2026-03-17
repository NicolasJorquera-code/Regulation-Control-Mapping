"""Excel ingest for existing control populations.

Parses uploaded Excel files (one sheet per section) into a list of
FinalControlRecord models. Handles type coercion, missing columns,
and string-encoded lists.
"""

from __future__ import annotations

import ast
import logging
from pathlib import Path
from typing import Any

import openpyxl

from controlnexus.core.state import FinalControlRecord

logger = logging.getLogger(__name__)

# 19 columns expected in each sheet (order matters for positional fallback)
EXPECTED_COLUMNS = [
    "control_id",
    "hierarchy_id",
    "leaf_name",
    "full_description",
    "selected_level_1",
    "selected_level_2",
    "business_unit_id",
    "business_unit_name",
    "who",
    "what",
    "when",
    "frequency",
    "where",
    "why",
    "quality_rating",
    "validator_passed",
    "validator_retries",
    "validator_failures",
    "evidence",
]


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ("true", "1", "yes")
    return bool(value) if value is not None else False


def _coerce_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def _parse_failures(value: Any) -> list[str]:
    """Parse validator_failures from string '[]' or '["X","Y"]' to list."""
    if isinstance(value, list):
        return value
    if not value or str(value).strip() in ("", "[]", "None"):
        return []
    try:
        parsed = ast.literal_eval(str(value))
        if isinstance(parsed, list):
            return [str(x) for x in parsed]
    except (ValueError, SyntaxError):
        pass
    return [str(value)]


def ingest_excel(path: Path | str) -> list[FinalControlRecord]:
    """Parse an Excel file into a list of FinalControlRecord.

    Reads all sheets whose names start with 'section_' (case-insensitive).
    Returns a flat list of all parsed control records.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records: list[FinalControlRecord] = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.lower().startswith("section_"):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Build column mapping from header row
        header_row = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
        col_map: dict[str, int] = {}
        for col_name in EXPECTED_COLUMNS:
            if col_name in header_row:
                col_map[col_name] = header_row.index(col_name)

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                def get(col: str, default: Any = "") -> Any:
                    idx = col_map.get(col)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return val if val is not None else default
                    return default

                record = FinalControlRecord(
                    control_id=str(get("control_id", "")),
                    hierarchy_id=str(get("hierarchy_id", "")),
                    leaf_name=str(get("leaf_name", "")),
                    full_description=str(get("full_description", "")),
                    selected_level_1=str(get("selected_level_1", "Unspecified")),
                    selected_level_2=str(get("selected_level_2", "")),
                    control_type=str(get("selected_level_2", "")),
                    business_unit_id=str(get("business_unit_id", "BU-UNSPECIFIED")),
                    business_unit_name=str(get("business_unit_name", "Unspecified")),
                    who=str(get("who", "")),
                    what=str(get("what", "")),
                    when=str(get("when", "")),
                    frequency=str(get("frequency", "Other")),
                    where=str(get("where", "")),
                    why=str(get("why", "")),
                    quality_rating=str(get("quality_rating", "Satisfactory")),
                    validator_passed=_coerce_bool(get("validator_passed", True)),
                    validator_retries=_coerce_int(get("validator_retries", 0)),
                    validator_failures=_parse_failures(get("validator_failures", "[]")),
                    evidence=str(get("evidence", "")),
                )
                records.append(record)
            except Exception:
                logger.warning("Skipping row %d in sheet %s", row_idx, sheet_name, exc_info=True)

    wb.close()
    logger.info("Ingested %d control records from %s", len(records), path.name)
    return records
