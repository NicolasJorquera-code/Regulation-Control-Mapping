"""Excel export for Risk Inventory Builder runs."""

from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from controlnexus.risk_inventory.models import ReviewDecision, RiskInventoryRecord, RiskInventoryRun, RiskInventoryWorkspace
from controlnexus.risk_inventory.services import (
    build_control_gaps,
    build_synthetic_control_recommendations,
)

RATING_FILLS = {
    "Low": "C6EBD6",
    "Medium": "FDDC69",
    "High": "FFB3B8",
    "Critical": "DA1E28",
    "Strong": "C6EBD6",
    "Satisfactory": "D0E2FF",
    "Improvement Needed": "FDDC69",
    "Inadequate": "FFB3B8",
}

DROPDOWNS = {
    "Does Risk Materialize?": ["Yes", "No"],
    "Control Design Effectiveness": ["Strong", "Satisfactory", "Improvement Needed", "Inadequate"],
    "Control Operating Effectiveness": ["Strong", "Satisfactory", "Improvement Needed", "Inadequate"],
    "Control Environment Rating": ["Strong", "Satisfactory", "Improvement Needed", "Inadequate"],
    "Management Response": ["accept", "mitigate", "monitor", "escalate"],
    "Required Validation Level": [
        "1LOD Process Owner",
        "Process Owner Challenge",
        "Specialist Owner Validation",
        "BU Head and 2LOD Challenge",
        "Executive Risk Committee",
    ],
    "Challenged Field": [
        "Risk Statement",
        "Impact",
        "Frequency",
        "Control Mapping",
        "Residual Risk",
        "KRI",
        "Evidence",
        "Management Response",
    ],
    "Evidence Sufficiency": ["Sufficient", "Needs More Evidence", "Not Supported"],
    "Final Decision": ["Approve", "Challenge", "Reject", "Escalate"],
    "Review Status": ["Not Started", "Pending Review", "Challenged", "Approved"],
    "Approval Status": ["Draft", "Approved", "Rejected"],
    "Checklist Status": ["Not Started", "In Review", "Ready to Approve", "Needs Evidence", "Escalated"],
    "Checklist Item Status": ["Not Started", "Complete", "Needs Evidence", "Challenge", "N/A"],
}


def export_risk_inventory_to_excel(run: RiskInventoryRun, output_path: Path | str) -> Path:
    """Write a multi-sheet risk inventory workbook to disk."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_risk_inventory_workbook(run)
    workbook.save(output_path)
    return output_path


def risk_inventory_excel_bytes(run: RiskInventoryRun) -> bytes:
    """Return a workbook as XLSX bytes for Streamlit download buttons."""
    buffer = BytesIO()
    build_risk_inventory_workbook(run).save(buffer)
    return buffer.getvalue()


def risk_inventory_workspace_excel_bytes(
    workspace: RiskInventoryWorkspace,
    review_decisions: list[ReviewDecision] | None = None,
) -> bytes:
    """Return a workspace executive demo artifact as XLSX bytes."""
    buffer = BytesIO()
    build_risk_inventory_workspace_workbook(workspace, review_decisions).save(buffer)
    return buffer.getvalue()


def risk_inventory_review_excel_bytes(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None = None,
    review_decisions: list[ReviewDecision] | None = None,
) -> bytes:
    """Return a Human-in-the-Loop review workbook as XLSX bytes."""
    buffer = BytesIO()
    build_hitl_review_workbook(run, workspace, review_decisions).save(buffer)
    return buffer.getvalue()


def build_hitl_review_workbook(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None = None,
    review_decisions: list[ReviewDecision] | None = None,
) -> openpyxl.Workbook:
    """Build a focused reviewer workbook with checklist, rationale, and decision log."""
    review_decisions = review_decisions or []
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "HITL Cover"
    _write_hitl_cover_sheet(cover, run, workspace)
    _write_rows(wb.create_sheet("Validation Queue"), _hitl_validation_queue_rows(run, workspace, review_decisions))
    _write_rows(wb.create_sheet("Review Checklist"), _hitl_checklist_rows(run, workspace))
    _write_rows(wb.create_sheet("Rationale Breakdown"), _hitl_rationale_rows(run, workspace))
    _write_rows(wb.create_sheet("Control Suggestions"), _hitl_control_suggestion_rows(run, workspace))
    _write_rows(wb.create_sheet("Evidence KRI Trace"), _hitl_evidence_kri_rows(run, workspace))
    _write_rows(wb.create_sheet("Reviewer Decision Log"), _hitl_decision_rows(run, workspace, review_decisions))
    for sheet in wb.worksheets:
        _format_sheet(sheet)
    _format_hitl_cover_sheet(cover)
    return wb


def build_risk_inventory_workspace_workbook(
    workspace: RiskInventoryWorkspace,
    review_decisions: list[ReviewDecision] | None = None,
) -> openpyxl.Workbook:
    """Build the polished workspace demo artifact."""
    review_decisions = review_decisions or []
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "Cover"
    _write_cover_sheet(cover, workspace)
    _write_rows(wb.create_sheet("Executive Summary"), _workspace_summary_rows(workspace))
    _write_rows(wb.create_sheet("BU Risk Heatmap"), _workspace_heatmap_rows(workspace))
    _write_rows(wb.create_sheet("BU Risk Breakdown"), _workspace_bu_breakdown_rows(workspace))
    _write_rows(wb.create_sheet("Process Risk Inventory"), _workspace_process_inventory_rows(workspace))
    _write_rows(wb.create_sheet("Risk Detail Dossier"), _workspace_risk_dossier_rows(workspace))
    _write_rows(wb.create_sheet("Control Gap Summary"), _workspace_control_gap_rows(workspace))
    _write_rows(wb.create_sheet("Synthetic Control Recs"), _workspace_synthetic_rows(workspace))
    _write_rows(wb.create_sheet("KRI Dashboard"), _workspace_kri_rows(workspace))
    _write_rows(wb.create_sheet("Review & Challenge"), _workspace_review_rows(workspace, review_decisions))
    _write_rows(wb.create_sheet("Reviewer Decision Log"), _workspace_review_decision_rows(workspace, review_decisions))
    _write_rows(wb.create_sheet("Source Trace"), _workspace_source_trace_rows(workspace))
    _write_rows(wb.create_sheet("Config Snapshot"), _workspace_config_rows(workspace))
    for sheet in wb.worksheets:
        _format_sheet(sheet)
    _format_cover_sheet(cover)
    _apply_heatmap_fills(wb["BU Risk Heatmap"])
    return wb


def _workspace_records(workspace: RiskInventoryWorkspace) -> list[RiskInventoryRecord]:
    return [record for run in workspace.runs for record in run.records]


def _process_lookup(workspace: RiskInventoryWorkspace) -> dict[str, Any]:
    return {process.process_id: process for process in workspace.processes}


def _bu_lookup(workspace: RiskInventoryWorkspace) -> dict[str, Any]:
    return {bu.bu_id: bu for bu in workspace.business_units}


def _business_unit_for_record(workspace: RiskInventoryWorkspace, record: RiskInventoryRecord) -> str:
    process = _process_lookup(workspace).get(record.process_id)
    bu = _bu_lookup(workspace).get(process.bu_id) if process else None
    return bu.bu_name if bu else ""


def _validation_for_record(workspace: RiskInventoryWorkspace, record: RiskInventoryRecord) -> dict[str, str]:
    process = _process_lookup(workspace).get(record.process_id)
    bu = _bu_lookup(workspace).get(process.bu_id) if process else None
    residual = record.residual_risk.residual_rating.value
    taxonomy = f"{record.taxonomy_node.level_1_category} {record.taxonomy_node.level_2_category}".lower()
    gaps = build_control_gaps(record)
    if residual == "Critical":
        level = "Executive Risk Committee"
        reviewer = "BU Head + 2LOD Risk Executive"
        basis = "Critical residual risk requires executive acceptance or remediation commitment."
    elif residual == "High":
        level = "BU Head and 2LOD Challenge"
        reviewer = f"{bu.head if bu else 'BU Head'} + 2LOD Operational Risk"
        basis = "High residual risk requires business ownership and independent challenge."
    elif any(term in taxonomy for term in ("cyber", "privacy", "data", "technology")):
        level = "Specialist Owner Validation"
        reviewer = "Technology/Data Owner + 2LOD Risk Partner"
        basis = "Specialist domain validation is needed for cyber, privacy, data, or technology exposure."
    elif gaps:
        level = "Process Owner Challenge"
        reviewer = process.owner if process else "Business Process Owner"
        basis = "Mapped coverage gaps require process-owner challenge before approval."
    else:
        level = "1LOD Process Owner"
        reviewer = process.owner if process else "Business Process Owner"
        basis = "Residual exposure is within normal process-owner approval authority."
    return {
        "Required Validation Level": level,
        "Required Reviewer": reviewer,
        "Validation Basis": basis,
        "Escalation Path": "1LOD Process Owner -> BU Head -> 2LOD Risk -> Executive Risk Committee",
    }


def _write_cover_sheet(ws: Worksheet, workspace: RiskInventoryWorkspace) -> None:
    records = _workspace_records(workspace)
    high_plus = sum(record.residual_risk.residual_rating.value in {"High", "Critical"} for record in records)
    controls = len(
        {
            mapping.control_id
            for record in records
            for mapping in record.control_mappings
            if mapping.control_id
        }
    )
    control_links = sum(len(record.control_mappings) for record in records)
    gaps = sum(len(build_control_gaps(record)) for record in records)
    ws["A1"] = "Risk Inventory Builder"
    ws["A2"] = "Executive Demo Artifact"
    ws["A4"] = "Institution"
    ws["B4"] = workspace.bank_name
    ws["A5"] = "Scope"
    ws["B5"] = "Focused payment exception workspace"
    ws["A6"] = "Generated"
    ws["B6"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A8"] = "Business Units"
    ws["B8"] = len(workspace.business_units)
    ws["C8"] = "Processes"
    ws["D8"] = len(workspace.processes)
    ws["E8"] = "Risk Records"
    ws["F8"] = len(records)
    ws["A9"] = "Mapped Controls"
    ws["B9"] = controls
    ws["C9"] = "Risk-Control Links"
    ws["D9"] = control_links
    ws["E9"] = "Open Control Gaps"
    ws["F9"] = gaps
    ws["A10"] = "High+ Residual"
    ws["B10"] = high_plus
    ws["A11"] = "Narrative"
    ws["B11"] = (
        "This workbook mirrors the Streamlit executive workbench: portfolio risk differences, "
        "process-by-process inventory, control gaps, KRI thresholds, review decisions, source trace, "
        "and configuration evidence."
    )


def _write_hitl_cover_sheet(
    ws: Worksheet,
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
) -> None:
    records = list(run.records)
    gaps = sum(len(build_control_gaps(record)) for record in records)
    high_plus = sum(record.residual_risk.residual_rating.value in {"High", "Critical"} for record in records)
    ws["A1"] = "Risk Inventory Builder"
    ws["A2"] = "Human-in-the-Loop Review Workbook"
    ws["A4"] = "Process"
    ws["B4"] = run.input_context.process_name
    ws["A5"] = "Business Unit"
    ws["B5"] = _run_business_unit_name(run, workspace)
    ws["A6"] = "Run ID"
    ws["B6"] = run.run_id
    ws["A7"] = "Generated"
    ws["B7"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A9"] = "Risk Records"
    ws["B9"] = len(records)
    ws["C9"] = "High+ Residual"
    ws["D9"] = high_plus
    ws["E9"] = "Control Gaps"
    ws["F9"] = gaps
    ws["A11"] = "Reviewer Workflow"
    ws["B11"] = (
        "Use the Validation Queue to prioritize review, the Checklist to record evidence-based challenge, "
        "the Rationale Breakdown to inspect scoring support, and the Decision Log to capture final approval."
    )


def _run_business_unit_name(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None = None,
) -> str:
    if workspace and run.records:
        return _business_unit_for_record(workspace, run.records[0]) or run.input_context.business_unit
    return run.input_context.business_unit


def _review_business_unit_for_record(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
    record: RiskInventoryRecord,
) -> str:
    return _business_unit_for_record(workspace, record) if workspace else run.input_context.business_unit


def _validation_for_review_record(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
    record: RiskInventoryRecord,
) -> dict[str, str]:
    if workspace:
        return _validation_for_record(workspace, record)
    residual = record.residual_risk.residual_rating.value
    taxonomy = f"{record.taxonomy_node.level_1_category} {record.taxonomy_node.level_2_category}".lower()
    gaps = build_control_gaps(record)
    if residual == "Critical":
        level = "Executive Risk Committee"
        reviewer = "BU Head + 2LOD Risk Executive"
        basis = "Critical residual risk requires executive acceptance or remediation commitment."
    elif residual == "High":
        level = "BU Head and 2LOD Challenge"
        reviewer = "BU Head + 2LOD Operational Risk"
        basis = "High residual risk requires business ownership and independent challenge."
    elif any(term in taxonomy for term in ("cyber", "privacy", "data", "technology")):
        level = "Specialist Owner Validation"
        reviewer = "Technology/Data Owner + 2LOD Risk Partner"
        basis = "Specialist domain validation is needed for cyber, privacy, data, or technology exposure."
    elif gaps:
        level = "Process Owner Challenge"
        reviewer = "Business Process Owner"
        basis = "Mapped coverage gaps require process-owner challenge before approval."
    else:
        level = "1LOD Process Owner"
        reviewer = "Business Process Owner"
        basis = "Residual exposure is within normal process-owner approval authority."
    return {
        "Required Validation Level": level,
        "Required Reviewer": reviewer,
        "Validation Basis": basis,
        "Escalation Path": "1LOD Process Owner -> BU Head -> 2LOD Risk -> Executive Risk Committee",
    }


def _decision_by_risk(review_decisions: list[ReviewDecision]) -> dict[str, ReviewDecision]:
    return {decision.risk_id: decision for decision in review_decisions}


def _review_status_value(record: RiskInventoryRecord, decision: ReviewDecision | None) -> str:
    if decision:
        return decision.review_status.value
    if record.review_challenges:
        return record.review_challenges[0].review_status.value
    return "Pending Review"


def _approval_status_value(record: RiskInventoryRecord, decision: ReviewDecision | None) -> str:
    if decision:
        return decision.approval_status.value
    if record.review_challenges:
        return record.review_challenges[0].approval_status.value
    return "Draft"


def _evidence_sufficiency(record: RiskInventoryRecord) -> str:
    return "Sufficient" if record.evidence_references else "Needs More Evidence"


def _open_issue_count(
    workspace: RiskInventoryWorkspace | None,
    record: RiskInventoryRecord,
) -> int:
    if not workspace:
        return sum(len(mapping.open_issues) for mapping in record.control_mappings)
    control_ids = {mapping.control_id for mapping in record.control_mappings}
    return sum(
        (
            issue.risk_id == record.risk_id
            or issue.process_id == record.process_id
            or bool(issue.control_id and issue.control_id in control_ids)
        )
        and issue.status.lower() not in {"closed", "resolved"}
        for issue in workspace.issues
    )


def _suggested_review_decision(record: RiskInventoryRecord, gaps: list[Any]) -> str:
    residual = record.residual_risk.residual_rating.value
    if residual == "Critical":
        return "Escalate"
    if gaps or residual == "High":
        return "Challenge"
    return "Approve"


def _hitl_validation_queue_rows(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
    review_decisions: list[ReviewDecision],
) -> list[dict[str, Any]]:
    decisions = _decision_by_risk(review_decisions)
    rows = []
    for record in run.records:
        decision = decisions.get(record.risk_id)
        validation = _validation_for_review_record(run, workspace, record)
        gaps = build_control_gaps(record)
        rows.append(
            {
                "Risk ID": record.risk_id,
                "Business Unit": _review_business_unit_for_record(run, workspace, record),
                "Process": record.process_name,
                "Level 1 Risk Category": record.taxonomy_node.level_1_category,
                "Level 2 Risk Category": record.taxonomy_node.level_2_category,
                "Residual Risk": record.residual_risk.residual_label,
                "Required Validation Level": validation["Required Validation Level"],
                "Required Reviewer": validation["Required Reviewer"],
                "Evidence Sufficiency": _evidence_sufficiency(record),
                "Control Gap Count": len(gaps),
                "Open Issue Count": _open_issue_count(workspace, record),
                "Final Decision": decision.final_approved_value if decision else _suggested_review_decision(record, gaps),
                "Checklist Status": "Not Started",
                "Reviewer": decision.reviewer if decision else validation["Required Reviewer"],
                "Review Status": _review_status_value(record, decision),
                "Approval Status": _approval_status_value(record, decision),
                "Reviewer Rationale": decision.reviewer_rationale if decision else validation["Validation Basis"],
                "Challenge Comments": decision.challenge_comments if decision else "",
            }
        )
    return rows


def _hitl_checklist_rows(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
) -> list[dict[str, Any]]:
    rows = []
    for record in run.records:
        validation = _validation_for_review_record(run, workspace, record)
        gaps = build_control_gaps(record)
        checklist = [
            (
                "Risk statement",
                "Does the risk statement match the process, event, causes, and consequences?",
                "Confirm wording or challenge ambiguity.",
                record.risk_statement.risk_description,
            ),
            (
                "Impact and frequency",
                "Are impact and frequency scores supported by metrics and rationale?",
                "Confirm scoring or document override.",
                f"Impact {int(record.impact_assessment.overall_impact_score)}; frequency {int(record.likelihood_assessment.likelihood_score)}.",
            ),
            (
                "Control coverage",
                "Do mapped controls address the root causes and residual exposure?",
                "Challenge partial coverage where gaps remain.",
                f"{len(record.control_mappings)} controls; {len(gaps)} detected gaps.",
            ),
            (
                "Evidence",
                "Is cited evidence current, relevant, and sufficient for the rating?",
                _evidence_sufficiency(record),
                "\n".join(ref.description for ref in record.evidence_references) or "No evidence references cited.",
            ),
            (
                "KRIs and appetite",
                "Are KRIs and appetite thresholds appropriate for this risk type?",
                "Confirm threshold fit or request additional monitoring.",
                record.risk_appetite.status if record.risk_appetite else "No risk appetite status attached.",
            ),
            (
                "Management response",
                "Is the recommended action proportionate to residual risk and due-date urgency?",
                record.residual_risk.management_response.response_type.value.title(),
                record.residual_risk.management_response.recommended_action,
            ),
            (
                "Approval authority",
                "Is the final decision routed to the required reviewer?",
                validation["Required Validation Level"],
                validation["Validation Basis"],
            ),
        ]
        for area, question, suggested_response, context in checklist:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Checklist Area": area,
                    "Review Question": question,
                    "Suggested Response": suggested_response,
                    "Evidence / Context": context,
                    "Reviewer Notes": "",
                    "Checklist Item Status": "Not Started",
                }
            )
    return rows


def _hitl_rationale_rows(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
) -> list[dict[str, Any]]:
    rows = []
    for record in run.records:
        validation = _validation_for_review_record(run, workspace, record)
        review = record.review_challenges[0] if record.review_challenges else None
        rows.append(
            {
                "Risk ID": record.risk_id,
                "Business Unit": _review_business_unit_for_record(run, workspace, record),
                "Level 2 Risk Category": record.taxonomy_node.level_2_category,
                "Risk Statement": record.risk_statement.risk_description,
                "Root Causes": "\n".join(record.risk_statement.causes),
                "Impact Rationale": record.impact_assessment.overall_impact_rationale,
                "Frequency Rationale": record.likelihood_assessment.rationale,
                "Residual Rationale": record.residual_risk.rationale,
                "Challenged Fields": ", ".join(review.challenged_fields) if review else "",
                "AI Original Value": _cell_value(review.ai_original_value) if review else "",
                "Reviewer Adjusted Value": _cell_value(review.reviewer_adjusted_value) if review else "",
                "Suggested Challenge Prompt": validation["Validation Basis"],
            }
        )
    return rows


def _hitl_control_suggestion_rows(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
) -> list[dict[str, Any]]:
    rows = []
    for record in run.records:
        gaps = build_control_gaps(record)
        if not gaps:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Row Type": "Coverage Note",
                    "Priority": "Low",
                    "Gap / Recommendation": "No material gap identified.",
                    "Suggested Owner": record.residual_risk.management_response.owner,
                    "Expected Evidence": "Maintain current evidence and periodic review cadence.",
                    "Reviewer Action": "Approve or add monitoring note.",
                }
            )
        for gap in gaps:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Row Type": "Detected Gap",
                    "Priority": gap.severity,
                    "Gap / Recommendation": gap.description,
                    "Suggested Owner": record.residual_risk.management_response.owner,
                    "Expected Evidence": "\n".join(gap.existing_control_ids),
                    "Reviewer Action": gap.recommendation,
                }
            )
        recommendations = (
            record.synthetic_control_recommendations
            or build_synthetic_control_recommendations(record, workspace)
        )
        for recommendation in recommendations:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Row Type": "Suggested Control",
                    "Priority": recommendation.priority,
                    "Gap / Recommendation": recommendation.control_statement,
                    "Suggested Owner": recommendation.suggested_owner,
                    "Expected Evidence": recommendation.expected_evidence,
                    "Reviewer Action": recommendation.rationale,
                }
            )
    return rows


def _hitl_evidence_kri_rows(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
) -> list[dict[str, Any]]:
    rows = []
    for record in run.records:
        for reference in record.evidence_references:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Trace Type": "Evidence",
                    "Trace ID": reference.evidence_id,
                    "Name": reference.description,
                    "Source / Owner": reference.source,
                    "Review Use": reference.evidence_type,
                    "Threshold / Detail": "Supports rating and reviewer challenge.",
                }
            )
        kris = workspace.kris_for_taxonomy(record.taxonomy_node.id) if workspace else []
        if not kris:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Trace Type": "KRI",
                    "Trace ID": "",
                    "Name": "No configured KRI linked",
                    "Source / Owner": "",
                    "Review Use": "Reviewer should confirm whether monitoring is needed.",
                    "Threshold / Detail": "",
                }
            )
        for kri in kris:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Trace Type": "KRI",
                    "Trace ID": kri.kri_id,
                    "Name": kri.kri_name,
                    "Source / Owner": kri.owner or kri.data_source,
                    "Review Use": kri.rationale,
                    "Threshold / Detail": f"Green: {kri.thresholds.green}; Amber: {kri.thresholds.amber}; Red: {kri.thresholds.red}",
                }
            )
    return rows


def _hitl_decision_rows(
    run: RiskInventoryRun,
    workspace: RiskInventoryWorkspace | None,
    review_decisions: list[ReviewDecision],
) -> list[dict[str, Any]]:
    decisions = _decision_by_risk(review_decisions)
    rows = []
    for record in run.records:
        decision = decisions.get(record.risk_id)
        validation = _validation_for_review_record(run, workspace, record)
        rows.append(
            {
                "Risk ID": record.risk_id,
                "Required Reviewer": validation["Required Reviewer"],
                "Reviewer": decision.reviewer if decision else "",
                "Review Status": _review_status_value(record, decision),
                "Approval Status": _approval_status_value(record, decision),
                "Final Decision": decision.final_approved_value if decision else "",
                "Reviewer Adjusted Value": decision.reviewer_adjusted_value if decision else "",
                "Reviewer Rationale": decision.reviewer_rationale if decision else "",
                "Challenge Comments": decision.challenge_comments if decision else "",
                "Decision Timestamp": decision.decided_at if decision else "",
            }
        )
    return rows


def _workspace_summary_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    rows = []
    for bu in workspace.business_units:
        records = []
        for process in workspace.processes_for_bu(bu.bu_id):
            run = workspace.run_for_process(process.process_id)
            records.extend(run.records if run else [])
        categories = Counter(record.taxonomy_node.level_1_category for record in records)
        rows.append(
            {
                "Business Unit": bu.bu_name,
                "BU Head": bu.head,
                "Processes": len(workspace.processes_for_bu(bu.bu_id)),
                "Risk Records": len(records),
                "Dominant Risk Category": categories.most_common(1)[0][0] if categories else "",
                "High+ Residual": sum(r.residual_risk.residual_rating.value in {"High", "Critical"} for r in records),
                "Control Gaps": sum(len(build_control_gaps(r)) for r in records),
                "Risk Profile Narrative": bu.risk_profile_summary,
            }
        )
    return rows


def _workspace_heatmap_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    rows = []
    for bu in workspace.business_units:
        records = []
        for process in workspace.processes_for_bu(bu.bu_id):
            run = workspace.run_for_process(process.process_id)
            records.extend(run.records if run else [])
        for category in sorted({record.taxonomy_node.level_1_category for record in _workspace_records(workspace)}):
            scoped = [record for record in records if record.taxonomy_node.level_1_category == category]
            rows.append(
                {
                    "Business Unit": bu.bu_name,
                    "Enterprise Risk Category": category,
                    "Risk Records": len(scoped),
                    "High+ Residual": sum(r.residual_risk.residual_rating.value in {"High", "Critical"} for r in scoped),
                    "Control Gaps": sum(len(build_control_gaps(r)) for r in scoped),
                    "Heat": _heat_value(scoped),
                }
            )
    return rows


def _heat_value(records: list[RiskInventoryRecord]) -> str:
    if not records:
        return "None"
    high_plus = sum(record.residual_risk.residual_rating.value in {"High", "Critical"} for record in records)
    if high_plus >= 2:
        return "High"
    if high_plus == 1:
        return "Medium"
    if len(records) >= 3:
        return "Elevated"
    return "Low"


def _workspace_bu_breakdown_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    rows = []
    for bu in workspace.business_units:
        records = []
        for process in workspace.processes_for_bu(bu.bu_id):
            run = workspace.run_for_process(process.process_id)
            records.extend(run.records if run else [])
        counts = Counter(record.taxonomy_node.level_1_category for record in records)
        for category, count in counts.most_common():
            scoped = [record for record in records if record.taxonomy_node.level_1_category == category]
            rows.append(
                {
                    "Business Unit": bu.bu_name,
                    "Risk Category": category,
                    "Risk Records": count,
                    "High+ Residual": sum(r.residual_risk.residual_rating.value in {"High", "Critical"} for r in scoped),
                    "Control Gap Count": sum(len(build_control_gaps(r)) for r in scoped),
                    "Primary Driver": bu.risk_profile_summary,
                }
            )
    return rows


def _workspace_process_inventory_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    process_lookup = _process_lookup(workspace)
    rows = []
    for run in workspace.runs:
        process = process_lookup.get(run.input_context.process_id)
        bu_name = _business_unit_for_record(workspace, run.records[0]) if run.records else run.input_context.business_unit
        for record in run.records:
            rows.append(
                {
                    "Business Unit": bu_name,
                    "Process ID": run.input_context.process_id,
                    "Process": run.input_context.process_name,
                    "Process Owner": process.owner if process else "",
                    "Risk ID": record.risk_id,
                    "L1 Risk Category": record.taxonomy_node.level_1_category,
                    "L2 Risk Category": record.taxonomy_node.level_2_category,
                    "Impact": int(record.impact_assessment.overall_impact_score),
                    "Frequency": int(record.likelihood_assessment.likelihood_score),
                    "Inherent Risk": record.inherent_risk.inherent_label,
                    "Residual Risk": record.residual_risk.residual_label,
                    "Mapped Controls": len(record.control_mappings),
                    "Control Gaps": len(build_control_gaps(record)),
                    "Management Response": record.residual_risk.management_response.response_type.value,
                }
            )
    return rows


def _workspace_risk_dossier_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    rows = []
    for record in _workspace_records(workspace):
        validation = _validation_for_record(workspace, record)
        rows.append(
            {
                "Business Unit": _business_unit_for_record(workspace, record),
                "Process": record.process_name,
                "Risk ID": record.risk_id,
                "Risk Statement": record.risk_statement.risk_description,
                "Root Causes": "\n".join(record.risk_statement.causes),
                "Impact Rationale": record.impact_assessment.overall_impact_rationale,
                "Frequency Rationale": record.likelihood_assessment.rationale,
                "Residual Rationale": record.residual_risk.rationale,
                "Required Validation Level": validation["Required Validation Level"],
                "Required Reviewer": validation["Required Reviewer"],
                "Validation Basis": validation["Validation Basis"],
            }
        )
    return rows


def _workspace_control_gap_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    return [
        {
            "Business Unit": _business_unit_for_record(workspace, record),
            "Process": record.process_name,
            "Risk ID": record.risk_id,
            "Risk Subcategory": record.taxonomy_node.level_2_category,
            "Gap Type": gap.gap_type,
            "Gap Severity": gap.severity,
            "Gap Description": gap.description,
            "Existing Controls": "\n".join(gap.existing_control_ids),
            "Recommendation": gap.recommendation,
        }
        for record in _workspace_records(workspace)
        for gap in (build_control_gaps(record) or [])
    ]


def _workspace_synthetic_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    return [
        {
            "Business Unit": _business_unit_for_record(workspace, record),
            "Process": record.process_name,
            "Risk ID": record.risk_id,
            "Recommendation ID": recommendation.recommendation_id,
            "Control Name": recommendation.control_name,
            "Control Type": recommendation.control_type,
            "Control Statement": recommendation.control_statement,
            "Rationale": recommendation.rationale,
            "Suggested Owner": recommendation.suggested_owner,
            "Frequency": recommendation.frequency,
            "Expected Evidence": recommendation.expected_evidence,
            "Priority": recommendation.priority,
        }
        for record in _workspace_records(workspace)
        for recommendation in (
            record.synthetic_control_recommendations
            or build_synthetic_control_recommendations(record, workspace)
        )
    ]


def _workspace_kri_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    l2_lookup = {node.id: node.level_2_category for node in workspace.risk_taxonomy_l2}
    return [
        {
            "KRI ID": kri.kri_id,
            "KRI": kri.kri_name,
            "Risk Subcategory": l2_lookup.get(kri.risk_taxonomy_id, kri.risk_taxonomy_id),
            "Definition": kri.metric_definition,
            "Formula": kri.formula,
            "Data Source": kri.data_source,
            "Frequency": kri.measurement_frequency,
            "Owner": kri.owner,
            "Green": kri.thresholds.green,
            "Amber": kri.thresholds.amber,
            "Red": kri.thresholds.red,
            "Escalation Path": kri.escalation_path,
        }
        for kri in workspace.kri_library
    ]


def _workspace_review_rows(
    workspace: RiskInventoryWorkspace,
    review_decisions: list[ReviewDecision],
) -> list[dict[str, Any]]:
    decision_by_risk = {decision.risk_id: decision for decision in review_decisions}
    rows = []
    for record in _workspace_records(workspace):
        validation = _validation_for_record(workspace, record)
        review = record.review_challenges[0] if record.review_challenges else None
        decision = decision_by_risk.get(record.risk_id)
        rows.append(
            {
                "Risk ID": record.risk_id,
                "Business Unit": _business_unit_for_record(workspace, record),
                "Process": record.process_name,
                "Residual Risk": record.residual_risk.residual_label,
                "Required Validation Level": validation["Required Validation Level"],
                "Required Reviewer": validation["Required Reviewer"],
                "Review Status": decision.review_status.value if decision else review.review_status.value if review else "Pending Review",
                "Approval Status": decision.approval_status.value if decision else review.approval_status.value if review else "Draft",
                "Challenged Field": ", ".join(review.challenged_fields) if review else "",
                "Evidence Sufficiency": "Sufficient" if record.evidence_references else "Needs More Evidence",
                "Final Decision": decision.final_approved_value if decision else "",
                "Reviewer Rationale": decision.reviewer_rationale if decision else review.reviewer_rationale if review else "",
                "Next Action": record.residual_risk.management_response.recommended_action,
            }
        )
    return rows


def _workspace_review_decision_rows(
    workspace: RiskInventoryWorkspace,
    review_decisions: list[ReviewDecision],
) -> list[dict[str, Any]]:
    decision_by_risk = {decision.risk_id: decision for decision in review_decisions}
    rows = []
    for record in _workspace_records(workspace):
        decision = decision_by_risk.get(record.risk_id)
        rows.append(
            {
                "Risk ID": record.risk_id,
                "Business Unit": _business_unit_for_record(workspace, record),
                "Process": record.process_name,
                "Reviewer": decision.reviewer if decision else "",
                "Review Status": decision.review_status.value if decision else "Pending Review",
                "Approval Status": decision.approval_status.value if decision else "Draft",
                "Challenged Field": "",
                "Challenge Comments": decision.challenge_comments if decision else "",
                "Adjusted Value": decision.reviewer_adjusted_value if decision else "",
                "Final Decision": decision.final_approved_value if decision else "",
                "Decision Timestamp": decision.decided_at if decision else "",
            }
        )
    return rows


def _workspace_source_trace_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    rows = []
    for run in workspace.runs:
        for source in run.input_context.source_documents:
            rows.append(
                {
                    "Trace Type": "Source Document",
                    "Business Unit": run.input_context.business_unit,
                    "Process": run.input_context.process_name,
                    "Reference": source,
                    "Detail": "Document/source used for risk inventory generation.",
                }
            )
        rows.extend(_scenario_basis_trace_rows(run, run.input_context.business_unit, run.input_context.process_name))
        for event in run.events:
            rows.append(
                {
                    "Trace Type": "Agent/Fallback Event",
                    "Business Unit": run.input_context.business_unit,
                    "Process": run.input_context.process_name,
                    "Reference": event.get("agent", ""),
                    "Detail": event.get("summary", ""),
                }
            )
    process_lookup = _process_lookup(workspace)
    for process in process_lookup.values():
        if process.apqc_crosswalk:
            rows.append(
                {
                    "Trace Type": "Optional APQC Crosswalk",
                    "Business Unit": _bu_lookup(workspace).get(process.bu_id).bu_name if _bu_lookup(workspace).get(process.bu_id) else "",
                    "Process": process.process_name,
                    "Reference": process.apqc_crosswalk.get("process_name", process.apqc_crosswalk.get("process_id", "")),
                    "Detail": process.apqc_crosswalk.get("rationale", ""),
                }
            )
    return rows


def _workspace_config_rows(workspace: RiskInventoryWorkspace) -> list[dict[str, Any]]:
    snapshot = workspace.runs[0].config_snapshot if workspace.runs else {}
    return _config_rows(snapshot)


def build_risk_inventory_workbook(run: RiskInventoryRun) -> openpyxl.Workbook:
    """Build the required multi-sheet workbook from a RiskInventoryRun."""
    mapped_controls = len(
        {
            mapping.control_id
            for record in run.records
            for mapping in record.control_mappings
            if mapping.control_id
        }
    )
    risk_control_links = sum(len(record.control_mappings) for record in run.records)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    _write_rows(
        ws,
        [
            {"Field": "Run ID", "Value": run.run_id},
            {"Field": "Process", "Value": run.input_context.process_name},
            {"Field": "Product", "Value": run.input_context.product},
            {"Field": "Business Unit", "Value": run.input_context.business_unit},
            {"Field": "Risk Records", "Value": len(run.records)},
            {"Field": "Mapped Controls", "Value": mapped_controls},
            {"Field": "Risk-Control Links", "Value": risk_control_links},
            {
                "Field": "High+ Residual Risks",
                "Value": sum(record.residual_risk.residual_rating.value in {"High", "Critical"} for record in run.records),
            },
            {"Field": "Headline", "Value": run.executive_summary.headline},
            {"Field": "Key Messages", "Value": "\n".join(run.executive_summary.key_messages)},
            {"Field": "Recommended Actions", "Value": "\n".join(run.executive_summary.recommended_actions)},
        ],
    )

    _write_rows(wb.create_sheet("BU Risk Heatmap"), _bu_heatmap_rows(run))
    _write_rows(wb.create_sheet("Risk Inventory"), [_inventory_row(record) for record in run.records])
    _write_rows(wb.create_sheet("Risk Statements and Root Causes"), _risk_statement_rows(run.records))
    _write_rows(wb.create_sheet("Control Coverage and Gaps"), _control_gap_rows(run.records))
    _write_rows(wb.create_sheet("Synthetic Control Recs"), _synthetic_control_rows(run.records))
    _write_rows(wb.create_sheet("Residual Risk Mgmt Actions"), _residual_action_rows(run.records))
    _write_rows(wb.create_sheet("KRI Program"), _kri_program_rows(run.records))
    _write_rows(wb.create_sheet("Source Trace Config Snapshot"), _source_trace_rows(run))

    # Legacy tabs retained for backward compatibility with existing demo/test flows.
    _write_rows(wb.create_sheet("Inherent Risk Assessment"), [_inherent_row(record) for record in run.records])
    _write_rows(wb.create_sheet("Control Mapping"), _control_mapping_rows(run.records))
    _write_rows(wb.create_sheet("Control Effectiveness"), _control_effectiveness_rows(run.records))
    _write_rows(wb.create_sheet("Residual Risk Assessment"), [_residual_row(record) for record in run.records])
    _write_rows(wb.create_sheet("Review and Challenge"), _review_rows(run.records))
    _write_rows(wb.create_sheet("Scoring Matrices"), _matrix_rows(run))
    _write_rows(wb.create_sheet("Configuration Snapshot"), _config_rows(run.config_snapshot))
    _write_rows(wb.create_sheet("Validation Findings"), [finding.model_dump() for finding in run.validation_findings])

    for sheet in wb.worksheets:
        _format_sheet(sheet)
    return wb


def _write_rows(ws: Worksheet, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["No records"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([_cell_value(row.get(header, "")) for header in headers])


def _inventory_row(record: RiskInventoryRecord) -> dict[str, Any]:
    impact_scores = {item.dimension.value: int(item.score) for item in record.impact_assessment.dimensions}
    return {
        "Risk ID": record.risk_id,
        "Process ID": record.process_id,
        "Process Name": record.process_name,
        "Product": record.product,
        "Level 1 Risk Category": record.taxonomy_node.level_1_category,
        "Level 2 Risk Category": record.taxonomy_node.level_2_category,
        "Risk Taxonomy Definition": record.taxonomy_node.definition,
        "Does Risk Materialize?": "Yes" if record.applicability.materializes else "No",
        "Materialization Type": record.applicability.materialization_type.value,
        "Applicability Rationale": record.applicability.rationale,
        "Risk Description": record.risk_statement.risk_description,
        "Exposure Metrics": "\n".join(f"{m.metric_name}: {m.metric_value}" for m in record.exposure_metrics),
        "Financial Impact": impact_scores.get("financial_impact", ""),
        "Regulatory Impact": impact_scores.get("regulatory_impact", ""),
        "Reputation Impact": impact_scores.get("reputational_impact", ""),
        "Overall Impact Score": int(record.impact_assessment.overall_impact_score),
        "Overall Impact Rationale": record.impact_assessment.overall_impact_rationale,
        "Overall Frequency Score": int(record.likelihood_assessment.likelihood_score),
        "Overall Frequency Rating": record.likelihood_assessment.likelihood_rating,
        "Overall Frequency Rationale": record.likelihood_assessment.rationale,
        "Overall Inherent Risk Score": record.inherent_risk.inherent_score,
        "Overall Inherent Risk Rating": record.inherent_risk.inherent_label,
        "Linked Controls": "\n".join(f"{m.control_id}: {m.control_name}" for m in record.control_mappings),
        "Control Design Effectiveness": record.control_environment.design_rating.value,
        "Control Operating Effectiveness": record.control_environment.operating_rating.value,
        "Control Environment Rating": record.control_environment.control_environment_rating.value,
        "Residual Risk Score": record.residual_risk.residual_score,
        "Residual Risk Rating": record.residual_risk.residual_label,
        "Residual Risk Rationale": record.residual_risk.rationale,
        "Management Response": record.residual_risk.management_response.response_type.value,
        "Recommended Action": record.residual_risk.management_response.recommended_action,
        "Review Status": record.review_challenges[0].review_status.value if record.review_challenges else "",
        "Reviewer Comments": record.review_challenges[0].challenge_comments if record.review_challenges else "",
        "Validation Flags": "\n".join(f.message for f in record.validation_findings),
        "Demo Record": "Yes" if record.demo_record else "No",
    }


def _inherent_row(record: RiskInventoryRecord) -> dict[str, Any]:
    return {
        "Risk ID": record.risk_id,
        "Level 2 Risk Category": record.taxonomy_node.level_2_category,
        "Impact Score": int(record.impact_assessment.overall_impact_score),
        "Frequency Score": int(record.likelihood_assessment.likelihood_score),
        "Inherent Score": record.inherent_risk.inherent_score,
        "Inherent Rating": record.inherent_risk.inherent_label,
        "Impact Rationale": record.impact_assessment.overall_impact_rationale,
        "Frequency Rationale": record.likelihood_assessment.rationale,
    }


def _residual_row(record: RiskInventoryRecord) -> dict[str, Any]:
    return {
        "Risk ID": record.risk_id,
        "Inherent Rating": record.inherent_risk.inherent_label,
        "Control Environment Rating": record.control_environment.control_environment_rating.value,
        "Residual Score": record.residual_risk.residual_score,
        "Residual Rating": record.residual_risk.residual_label,
        "Management Response": record.residual_risk.management_response.response_type.value,
        "Recommended Action": record.residual_risk.management_response.recommended_action,
        "Rationale": record.residual_risk.rationale,
    }


def _control_mapping_rows(records: list[RiskInventoryRecord]) -> list[dict[str, Any]]:
    rows = []
    for record in records:
        for mapping in record.control_mappings:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Level 2 Risk Category": record.taxonomy_node.level_2_category,
                    "Control ID": mapping.control_id,
                    "Control Name": mapping.control_name,
                    "Control Type": mapping.control_type,
                    "Coverage Assessment": mapping.coverage_assessment,
                    "Mapping Rationale": mapping.mitigation_rationale,
                }
            )
    return rows


def _control_effectiveness_rows(records: list[RiskInventoryRecord]) -> list[dict[str, Any]]:
    rows = []
    for record in records:
        for mapping in record.control_mappings:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Control ID": mapping.control_id,
                    "Control Name": mapping.control_name,
                    "Design Rating": mapping.design_effectiveness.rating.value if mapping.design_effectiveness else "",
                    "Design Rationale": mapping.design_effectiveness.rationale if mapping.design_effectiveness else "",
                    "Operating Rating": mapping.operating_effectiveness.rating.value if mapping.operating_effectiveness else "",
                    "Operating Rationale": mapping.operating_effectiveness.rationale if mapping.operating_effectiveness else "",
                    "Control Environment Rating": record.control_environment.control_environment_rating.value,
                }
            )
    return rows


def _bu_heatmap_rows(run: RiskInventoryRun) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[RiskInventoryRecord]] = {}
    for record in run.records:
        key = (
            run.input_context.business_unit or "Unassigned",
            record.taxonomy_node.level_1_category,
            record.taxonomy_node.level_2_category,
        )
        grouped.setdefault(key, []).append(record)

    rows = []
    for (business_unit, l1, l2), records in sorted(grouped.items()):
        rows.append(
            {
                "Business Unit": business_unit,
                "Process": run.input_context.process_name,
                "Level 1 Risk Category": l1,
                "Level 2 Risk Category": l2,
                "Risk Records": len(records),
                "High+ Inherent": sum(r.inherent_risk.inherent_rating.value in {"High", "Critical"} for r in records),
                "High+ Residual": sum(r.residual_risk.residual_rating.value in {"High", "Critical"} for r in records),
                "Mapped Controls": sum(len(r.control_mappings) for r in records),
                "Coverage Gaps": sum(len(build_control_gaps(r)) for r in records),
            }
        )
    return rows


def _risk_statement_rows(records: list[RiskInventoryRecord]) -> list[dict[str, Any]]:
    return [
        {
            "Risk ID": record.risk_id,
            "Process": record.process_name,
            "Level 1 Risk Category": record.taxonomy_node.level_1_category,
            "Level 2 Risk Category": record.taxonomy_node.level_2_category,
            "Risk Statement": record.risk_statement.risk_description,
            "Risk Event": record.risk_statement.risk_event,
            "Root Causes": "\n".join(record.risk_statement.causes),
            "Consequences": "\n".join(record.risk_statement.consequences),
            "Affected Stakeholders": "\n".join(record.risk_statement.affected_stakeholders),
            "Evidence References": "\n".join(
                f"{ref.evidence_id}: {ref.description}" for ref in record.evidence_references
            ),
        }
        for record in records
    ]


def _control_gap_rows(records: list[RiskInventoryRecord]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        gaps = build_control_gaps(record)
        if not gaps:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Process": record.process_name,
                    "Level 2 Risk Category": record.taxonomy_node.level_2_category,
                    "Gap Type": "No material gap",
                    "Gap Severity": "",
                    "Gap Description": "",
                    "Existing Controls": "\n".join(m.control_id for m in record.control_mappings),
                    "Recommendation": "Maintain current coverage and evidence discipline.",
                }
            )
            continue
        for gap in gaps:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Process": record.process_name,
                    "Level 2 Risk Category": record.taxonomy_node.level_2_category,
                    "Gap Type": gap.gap_type,
                    "Gap Severity": gap.severity,
                    "Gap Description": gap.description,
                    "Root Causes": "\n".join(gap.root_causes),
                    "Existing Controls": "\n".join(gap.existing_control_ids),
                    "Recommendation": gap.recommendation,
                }
            )
    return rows


def _synthetic_control_rows(records: list[RiskInventoryRecord]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        recommendations = (
            record.synthetic_control_recommendations
            or build_synthetic_control_recommendations(record)
        )
        for recommendation in recommendations:
            rows.append(
                {
                    "Recommendation ID": recommendation.recommendation_id,
                    "Risk ID": record.risk_id,
                    "Process": record.process_name,
                    "Level 2 Risk Category": record.taxonomy_node.level_2_category,
                    "Control Name": recommendation.control_name,
                    "Control Type": recommendation.control_type,
                    "Control Statement": recommendation.control_statement,
                    "Rationale": recommendation.rationale,
                    "Addressed Root Causes": "\n".join(recommendation.addressed_root_causes),
                    "Suggested Owner": recommendation.suggested_owner,
                    "Frequency": recommendation.frequency,
                    "Expected Evidence": recommendation.expected_evidence,
                    "Priority": recommendation.priority,
                }
            )
    return rows


def _residual_action_rows(records: list[RiskInventoryRecord]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        if record.action_plan:
            for action in record.action_plan:
                rows.append(
                    {
                        "Risk ID": record.risk_id,
                        "Process": record.process_name,
                        "Residual Risk Rating": record.residual_risk.residual_label,
                        "Management Response": record.residual_risk.management_response.response_type.value,
                        "Recommended Action": record.residual_risk.management_response.recommended_action,
                        "Action Item": action.action,
                        "Owner": action.owner,
                        "Due Date": action.due_date,
                        "Priority": action.priority,
                        "Status": action.status,
                    }
                )
        else:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Process": record.process_name,
                    "Residual Risk Rating": record.residual_risk.residual_label,
                    "Management Response": record.residual_risk.management_response.response_type.value,
                    "Recommended Action": record.residual_risk.management_response.recommended_action,
                    "Action Item": "",
                    "Owner": record.residual_risk.management_response.owner,
                    "Due Date": record.residual_risk.management_response.due_date,
                    "Priority": "",
                    "Status": "",
                }
            )
    return rows


def _kri_program_rows(records: list[RiskInventoryRecord]) -> list[dict[str, Any]]:
    return [
        {
            "Risk ID": record.risk_id,
            "Process": record.process_name,
            "Level 2 Risk Category": record.taxonomy_node.level_2_category,
            "Recommended KRI Family": _kri_family(record),
            "Primary Measurement": ", ".join(metric.metric_name for metric in record.exposure_metrics[:3]),
            "Residual Risk Rating": record.residual_risk.residual_label,
            "Cadence Guidance": "Weekly" if record.residual_risk.residual_rating.value in {"High", "Critical"} else "Monthly",
            "Escalation Trigger": "Amber trend or any red threshold breach requires business owner and 2LOD review.",
            "Threshold Design Note": "Calibrate green/amber/red thresholds with production history and risk appetite.",
        }
        for record in records
    ]


def _source_trace_rows(run: RiskInventoryRun) -> list[dict[str, Any]]:
    rows = [
        {
            "Trace Type": "Source Document",
            "Name": source,
            "Stage": "Data Intake",
            "Detail": run.input_context.process_name,
        }
        for source in run.input_context.source_documents
    ]
    rows.extend(
        {
            "Trace Type": "Scenario Basis",
            "Name": row["Reference"],
            "Stage": "Public Source Trace",
            "Detail": row["Detail"],
        }
        for row in _scenario_basis_trace_rows(run, "", "")
    )
    rows.extend(
        {
            "Trace Type": "Agent Event",
            "Name": event.get("agent", ""),
            "Stage": event.get("stage", ""),
            "Detail": event.get("summary", ""),
        }
        for event in run.events
    )
    rows.extend(
        {
            "Trace Type": "Run Manifest",
            "Name": key,
            "Stage": "Run Assembly",
            "Detail": _cell_value(value),
        }
        for key, value in run.run_manifest.items()
    )
    rows.extend(
        {
            "Trace Type": "Config Snapshot",
            "Name": row["Config Path"],
            "Stage": "Configuration",
            "Detail": row["Value"],
        }
        for row in _config_rows(run.config_snapshot)
    )
    return rows


def _scenario_basis_trace_rows(
    run: RiskInventoryRun,
    business_unit: str,
    process: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source in run.run_manifest.get("scenario_basis", []):
        if not isinstance(source, dict):
            continue
        title = str(source.get("title", "")).strip()
        url = str(source.get("url", "")).strip()
        relevance = str(source.get("relevance", "")).strip()
        rows.append(
            {
                "Trace Type": "Scenario Basis",
                "Business Unit": business_unit,
                "Process": process,
                "Reference": title or url,
                "Detail": f"{url} | {relevance}".strip(" |"),
            }
        )
    return rows


def _kri_family(record: RiskInventoryRecord) -> str:
    category = record.taxonomy_node.level_2_category.lower()
    if "data" in category or "reporting" in category:
        return "Data quality, timeliness, and exception aging"
    if "cyber" in category or "privacy" in category:
        return "Access, exception aging, and incident timeliness"
    if "third party" in category:
        return "Vendor SLA, concentration, and unresolved issue trend"
    if "fraud" in category:
        return "Exception volume, confirmed event rate, and investigation aging"
    return "Process volume, SLA breaches, and control exceptions"


def _review_rows(records: list[RiskInventoryRecord]) -> list[dict[str, Any]]:
    rows = []
    for record in records:
        for review in record.review_challenges:
            rows.append(
                {
                    "Risk ID": record.risk_id,
                    "Review Status": review.review_status.value,
                    "Reviewer": review.reviewer,
                    "Challenged Fields": ", ".join(review.challenged_fields),
                    "Challenge Comments": review.challenge_comments,
                    "Reviewer Rationale": review.reviewer_rationale,
                    "Approval Status": review.approval_status.value,
                }
            )
    return rows


def _matrix_rows(run: RiskInventoryRun) -> list[dict[str, Any]]:
    rows = []
    inherent = run.config_snapshot.get("inherent_risk_matrix", {}).get("matrix", {})
    for impact, likelihoods in inherent.items():
        for likelihood, result in likelihoods.items():
            rows.append({"Matrix": "Inherent", "Impact": impact, "Frequency/Environment": likelihood, **result})
    residual = run.config_snapshot.get("residual_risk_matrix", {}).get("matrix", {})
    for inherent_label, environments in residual.items():
        for environment, result in environments.items():
            rows.append(
                {
                    "Matrix": "Residual",
                    "Impact": inherent_label,
                    "Frequency/Environment": environment,
                    **result,
                }
            )
    return rows


def _config_rows(snapshot: dict[str, Any], prefix: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key, value in snapshot.items():
        path = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(_config_rows(value, path))
        else:
            rows.append({"Config Path": path, "Value": _cell_value(value)})
    return rows


def _cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    return value


def _format_sheet(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    if ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="161616")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            fill = RATING_FILLS.get(str(cell.value), "")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 14), 42)
    _add_dropdowns(ws)


def _format_cover_sheet(ws: Worksheet) -> None:
    ws.freeze_panes = None
    ws.auto_filter.ref = None
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws.merge_cells("B11:F13")
    ws["A1"].font = Font(size=22, bold=True, color="FFFFFF")
    ws["A2"].font = Font(size=14, bold=True, color="D0E2FF")
    ws["A1"].fill = PatternFill("solid", fgColor="161616")
    ws["A2"].fill = PatternFill("solid", fgColor="262626")
    ws["B11"].alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(4, 10):
        ws[f"A{row}"].font = Font(bold=True, color="525252")
        ws[f"C{row}"].font = Font(bold=True, color="525252")
        ws[f"E{row}"].font = Font(bold=True, color="525252")
    thin = Side(style="thin", color="C6C6C6")
    for row in ws.iter_rows(min_row=4, max_row=13, min_col=1, max_col=6):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ["A", "C", "E"]:
        ws.column_dimensions[col].width = 22
    for col in ["B", "D", "F"]:
        ws.column_dimensions[col].width = 24


def _format_hitl_cover_sheet(ws: Worksheet) -> None:
    ws.freeze_panes = None
    ws.auto_filter.ref = None
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws.merge_cells("B11:F13")
    ws["A1"].font = Font(size=22, bold=True, color="FFFFFF")
    ws["A2"].font = Font(size=14, bold=True, color="D0E2FF")
    ws["A1"].fill = PatternFill("solid", fgColor="161616")
    ws["A2"].fill = PatternFill("solid", fgColor="262626")
    ws["B11"].alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(4, 12):
        ws[f"A{row}"].font = Font(bold=True, color="525252")
        ws[f"C{row}"].font = Font(bold=True, color="525252")
        ws[f"E{row}"].font = Font(bold=True, color="525252")
    thin = Side(style="thin", color="C6C6C6")
    for row in ws.iter_rows(min_row=4, max_row=13, min_col=1, max_col=6):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ["A", "C", "E"]:
        ws.column_dimensions[col].width = 22
    for col in ["B", "D", "F"]:
        ws.column_dimensions[col].width = 28


def _apply_heatmap_fills(ws: Worksheet) -> None:
    heat_colors = {
        "None": "F4F4F4",
        "Low": "C6EBD6",
        "Elevated": "FFF1C7",
        "Medium": "FDDC69",
        "High": "FFB3B8",
    }
    headers = {str(cell.value): cell.column for cell in ws[1] if cell.value}
    heat_col = headers.get("Heat")
    if not heat_col:
        return
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=heat_col)
        color = heat_colors.get(str(cell.value))
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
    high_col = headers.get("High+ Residual")
    if high_col and ws.max_row > 1:
        col_letter = ws.cell(row=1, column=high_col).column_letter
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{ws.max_row}",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor="FFB3B8"),
            ),
        )


def _add_dropdowns(ws: Worksheet) -> None:
    headers = {str(cell.value): cell.column_letter for cell in ws[1] if cell.value}
    for header, values in DROPDOWNS.items():
        column = headers.get(header)
        if not column:
            continue
        validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"{column}2:{column}1048576")
