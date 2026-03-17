"""Parse APQC hierarchy data from Excel or CSV into HierarchyNode objects."""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import Any

import openpyxl

from controlnexus.core.state import HierarchyNode

logger = logging.getLogger(__name__)

# Section sheet names used in the APQC Excel template
SECTION_SHEETS = [f"{i}.0" for i in range(1, 14)]

# Column name mappings (handle case variations)
_COL_ALIASES: dict[str, str] = {
    "pcf id": "pcf_id",
    "hierarchy id": "hierarchy_id",
    "name": "name",
    "difference index": "difference_index",
    "change details": "change_details",
    "metrics available?": "metrics_available",
    "metrics available": "metrics_available",
}


def _normalize_col(header: str) -> str:
    """Map a raw column header to its internal field name."""
    return _COL_ALIASES.get(header.strip().lower(), header.strip().lower())


def _parse_hierarchy_id(hid: str) -> tuple[int, str, str | None]:
    """Derive depth, top_section, and parent_hierarchy_id from a hierarchy ID.

    Examples:
        "4.0"     -> (depth=2, top="4", parent=None)
        "4.1"     -> (depth=2, top="4", parent="4.0")
        "4.1.1"   -> (depth=3, top="4", parent="4.1")
        "4.1.1.1" -> (depth=4, top="4", parent="4.1.1")
    """
    parts = hid.split(".")
    top_section = parts[0]
    depth = len(parts)

    if depth <= 1:
        return depth, top_section, None

    # For "X.0" nodes (section roots), parent is None
    if depth == 2 and parts[1] == "0":
        return depth, top_section, None

    # For depth-2 non-root nodes (e.g., "4.1"), parent is the section root "4.0"
    if depth == 2:
        return depth, top_section, f"{top_section}.0"

    # Otherwise drop the last segment (e.g., "4.1.1" → "4.1")
    parent = ".".join(parts[:-1])
    return depth, top_section, parent


def _row_to_node(
    row: dict[str, Any],
    source_sheet: str = "",
    source_row: int = 0,
) -> HierarchyNode | None:
    """Convert a normalized row dict to a HierarchyNode (or None if invalid)."""
    hierarchy_id = str(row.get("hierarchy_id", "")).strip()
    if not hierarchy_id:
        return None

    name = str(row.get("name", "")).strip()
    if not name:
        return None

    depth, top_section, parent_hid = _parse_hierarchy_id(hierarchy_id)

    pcf_id_raw = row.get("pcf_id", "")
    pcf_id = str(int(pcf_id_raw)) if isinstance(pcf_id_raw, float) else str(pcf_id_raw or "")

    diff_idx_raw = row.get("difference_index", 0)
    try:
        difference_index = float(diff_idx_raw) if diff_idx_raw else 0.0
    except (ValueError, TypeError):
        difference_index = 0.0

    change_details = row.get("change_details") or None
    if change_details:
        change_details = str(change_details).strip() or None

    metrics_raw = row.get("metrics_available", "")
    metrics_available = str(metrics_raw).strip().upper() == "Y" if metrics_raw else None

    return HierarchyNode(
        pcf_id=pcf_id,
        hierarchy_id=hierarchy_id,
        name=name,
        depth=depth,
        top_section=top_section,
        is_leaf=False,  # set in post-processing
        parent_hierarchy_id=parent_hid,
        source_sheet=source_sheet,
        source_row=source_row,
        difference_index=difference_index,
        change_details=change_details,
        metrics_available=metrics_available,
    )


def _mark_leaves(nodes: list[HierarchyNode]) -> list[HierarchyNode]:
    """Mark leaf nodes (nodes that are not a parent of any other node)."""
    parent_ids = {n.parent_hierarchy_id for n in nodes if n.parent_hierarchy_id}
    result = []
    for node in nodes:
        if node.hierarchy_id in parent_ids:
            result.append(node)
        else:
            result.append(node.model_copy(update={"is_leaf": True}))
    return result


def _deduplicate(nodes: list[HierarchyNode]) -> list[HierarchyNode]:
    """Remove duplicate nodes by hierarchy_id, keeping the first occurrence."""
    seen: set[str] = set()
    unique: list[HierarchyNode] = []
    for node in nodes:
        if node.hierarchy_id not in seen:
            seen.add(node.hierarchy_id)
            unique.append(node)
    return unique


# -- Excel loading -------------------------------------------------------------


def _load_from_excel(path: Path) -> list[HierarchyNode]:
    """Load hierarchy nodes from an APQC Excel template."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    nodes: list[HierarchyNode] = []

    # Prefer the "Combined" sheet if it exists
    if "Combined" in wb.sheetnames:
        ws = wb["Combined"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [_normalize_col(str(h or "")) for h in rows[0]]
            for row_idx, row_values in enumerate(rows[1:], start=2):
                row_dict = dict(zip(headers, row_values))
                node = _row_to_node(row_dict, source_sheet="Combined", source_row=row_idx)
                if node:
                    nodes.append(node)
        logger.info("Loaded %d nodes from Combined sheet", len(nodes))
    else:
        # Fall back to per-section sheets
        for sheet_name in SECTION_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [_normalize_col(str(h or "")) for h in rows[0]]
            for row_idx, row_values in enumerate(rows[1:], start=2):
                row_dict = dict(zip(headers, row_values))
                node = _row_to_node(row_dict, source_sheet=sheet_name, source_row=row_idx)
                if node:
                    nodes.append(node)
        logger.info("Loaded %d nodes from %d section sheets", len(nodes), len(SECTION_SHEETS))

    wb.close()
    return nodes


# -- CSV loading ---------------------------------------------------------------


def _load_from_csv(path: Path) -> list[HierarchyNode]:
    """Load hierarchy nodes from a CSV file."""
    nodes: list[HierarchyNode] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, start=2):
            normalized = {_normalize_col(k): v for k, v in row.items()}
            node = _row_to_node(normalized, source_sheet="csv", source_row=row_idx)
            if node:
                nodes.append(node)
    logger.info("Loaded %d nodes from CSV %s", len(nodes), path.name)
    return nodes


def _load_from_csv_text(text: str) -> list[HierarchyNode]:
    """Load hierarchy nodes from CSV text content."""
    nodes: list[HierarchyNode] = []
    reader = csv.DictReader(io.StringIO(text))
    for row_idx, row in enumerate(reader, start=2):
        normalized = {_normalize_col(k): v for k, v in row.items()}
        node = _row_to_node(normalized, source_sheet="csv", source_row=row_idx)
        if node:
            nodes.append(node)
    return nodes


# -- Public API ----------------------------------------------------------------


def load_apqc_hierarchy(source: Path) -> list[HierarchyNode]:
    """Load and parse APQC hierarchy from an Excel (.xlsx) or CSV file.

    Args:
        source: Path to the APQC template file.

    Returns:
        List of HierarchyNode objects with leaf status computed.

    Raises:
        FileNotFoundError: If the source file does not exist.
        ValueError: If the file extension is not supported.
    """
    source = Path(source)
    if not source.exists():
        raise FileNotFoundError(f"APQC hierarchy file not found: {source}")

    ext = source.suffix.lower()
    if ext == ".xlsx":
        nodes = _load_from_excel(source)
    elif ext == ".csv":
        nodes = _load_from_csv(source)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Expected .xlsx or .csv")

    nodes = _deduplicate(nodes)
    nodes = _mark_leaves(nodes)
    logger.info(
        "APQC hierarchy loaded: %d nodes, %d leaves",
        len(nodes),
        sum(1 for n in nodes if n.is_leaf),
    )
    return nodes


def load_apqc_hierarchy_from_bytes(data: bytes, filename: str) -> list[HierarchyNode]:
    """Load APQC hierarchy from in-memory bytes (for Streamlit file uploader).

    Args:
        data: Raw file content bytes.
        filename: Original filename (used to detect format).

    Returns:
        List of HierarchyNode objects with leaf status computed.
    """
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            tmp.flush()
            nodes = _load_from_excel(Path(tmp.name))
        Path(tmp.name).unlink(missing_ok=True)
    elif ext == ".csv":
        text = data.decode("utf-8-sig")
        nodes = _load_from_csv_text(text)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Expected .xlsx or .csv")

    nodes = _deduplicate(nodes)
    nodes = _mark_leaves(nodes)
    return nodes
