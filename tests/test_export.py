"""Tests for controlnexus.export.excel."""

from __future__ import annotations

import openpyxl

from controlnexus.core.state import FinalControlRecord
from controlnexus.export.excel import EXPORT_COLUMNS, export_to_excel


def _make_control(**overrides) -> FinalControlRecord:
    base = {
        "control_id": "CTRL-001",
        "hierarchy_id": "4.1.1.1",
        "leaf_name": "Test",
        "full_description": "Test description",
        "selected_level_1": "Preventive",
        "selected_level_2": "Reconciliation",
        "who": "Analyst",
        "what": "Reviews",
        "when": "Monthly",
        "where": "GL System",
        "why": "Prevent risk",
    }
    base.update(overrides)
    return FinalControlRecord(**base)


class TestExportToExcel:
    def test_creates_file(self, tmp_path):
        records = [_make_control()]
        path = export_to_excel(records, tmp_path / "output.xlsx")
        assert path.exists()

    def test_correct_headers(self, tmp_path):
        records = [_make_control()]
        path = export_to_excel(records, tmp_path / "output.xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == EXPORT_COLUMNS

    def test_correct_row_count(self, tmp_path):
        records = [_make_control(control_id=f"CTRL-{i}") for i in range(5)]
        path = export_to_excel(records, tmp_path / "output.xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.max_row == 6  # 1 header + 5 data

    def test_round_trip(self, tmp_path):
        records = [_make_control(validator_failures=["VAGUE_WHEN"])]
        path = export_to_excel(records, tmp_path / "output.xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Check control_id in first data row
        assert ws.cell(row=2, column=1).value == "CTRL-001"

    def test_empty_records(self, tmp_path):
        path = export_to_excel([], tmp_path / "empty.xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.max_row == 1  # header only
