"""Tests for controlnexus.analysis.ingest."""

from __future__ import annotations


import openpyxl

from controlnexus.analysis.ingest import (
    _coerce_bool,
    _coerce_int,
    _parse_failures,
    ingest_excel,
)


class TestCoerceBool:
    def test_true_values(self):
        assert _coerce_bool(True) is True
        assert _coerce_bool("true") is True
        assert _coerce_bool("True") is True
        assert _coerce_bool("1") is True
        assert _coerce_bool("yes") is True

    def test_false_values(self):
        assert _coerce_bool(False) is False
        assert _coerce_bool("false") is False
        assert _coerce_bool("0") is False
        assert _coerce_bool(None) is False
        assert _coerce_bool("") is False


class TestCoerceInt:
    def test_valid(self):
        assert _coerce_int(5) == 5
        assert _coerce_int("3") == 3
        assert _coerce_int(0) == 0

    def test_invalid(self):
        assert _coerce_int(None) == 0
        assert _coerce_int("abc") == 0
        assert _coerce_int("", 99) == 99


class TestParseFailures:
    def test_empty_string(self):
        assert _parse_failures("[]") == []
        assert _parse_failures("") == []
        assert _parse_failures(None) == []

    def test_list_passthrough(self):
        assert _parse_failures(["A", "B"]) == ["A", "B"]

    def test_string_list(self):
        result = _parse_failures('["VAGUE_WHEN", "SPEC_MISMATCH"]')
        assert result == ["VAGUE_WHEN", "SPEC_MISMATCH"]

    def test_unparseable(self):
        result = _parse_failures("not a list")
        assert result == ["not a list"]


class TestIngestExcel:
    def test_ingest_basic(self, tmp_path):
        """Create a minimal Excel file and parse it."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "section_4_controls"

        # Header row
        headers = [
            "control_id",
            "hierarchy_id",
            "leaf_name",
            "full_description",
            "selected_level_1",
            "selected_level_2",
            "business_unit_id",
            "business_unit_name",
            "who",
            "what",
            "when",
            "frequency",
            "where",
            "why",
            "quality_rating",
            "validator_passed",
            "validator_retries",
            "validator_failures",
            "evidence",
        ]
        ws.append(headers)

        # Data row
        ws.append(
            [
                "CTRL-0401-REC-001",
                "4.1.1.1",
                "Test Leaf",
                "Monthly reconciliation of accounts to prevent discrepancies.",
                "Preventive",
                "Reconciliation",
                "BU-001",
                "Retail Banking",
                "Senior Accountant",
                "Reconciles accounts",
                "Monthly",
                "Monthly",
                "GL System",
                "Prevent discrepancies",
                "Strong",
                True,
                0,
                "[]",
                "Reconciliation report with sign-off",
            ]
        )

        path = tmp_path / "test.xlsx"
        wb.save(path)

        records = ingest_excel(path)
        assert len(records) == 1
        assert records[0].control_id == "CTRL-0401-REC-001"
        assert records[0].hierarchy_id == "4.1.1.1"
        assert records[0].validator_passed is True
        assert records[0].validator_failures == []

    def test_skips_non_section_sheets(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Categories"
        ws.append(["col1"])
        ws.append(["data"])

        path = tmp_path / "test.xlsx"
        wb.save(path)

        records = ingest_excel(path)
        assert len(records) == 0

    def test_multiple_rows(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "section_9_controls"

        headers = [
            "control_id",
            "hierarchy_id",
            "leaf_name",
            "full_description",
            "selected_level_1",
            "selected_level_2",
            "business_unit_id",
            "business_unit_name",
            "who",
            "what",
            "when",
            "frequency",
            "where",
            "why",
            "quality_rating",
            "validator_passed",
            "validator_retries",
            "validator_failures",
            "evidence",
        ]
        ws.append(headers)

        for i in range(5):
            ws.append(
                [
                    f"CTRL-090{i}",
                    f"9.1.{i}",
                    f"Leaf {i}",
                    "Description words " * 5,
                    "Preventive",
                    "Reconciliation",
                    "BU-001",
                    "Treasury",
                    "Analyst",
                    "Reviews",
                    "Monthly",
                    "Monthly",
                    "System",
                    "Prevent risk",
                    "Satisfactory",
                    True,
                    0,
                    "[]",
                    "Report",
                ]
            )

        path = tmp_path / "test.xlsx"
        wb.save(path)

        records = ingest_excel(path)
        assert len(records) == 5
