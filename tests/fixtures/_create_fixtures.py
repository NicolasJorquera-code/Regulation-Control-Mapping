"""One-time script to generate test fixture Excel files."""
import openpyxl
from pathlib import Path

FIXTURE_DIR = Path(__file__).parent

# ── sample_register.xlsx: standard header names, 2 sheets, 10 controls ──
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "section_1"
headers = [
    "Control ID", "Control Type", "Business Unit", "Section",
    "Placement", "Method", "Frequency", "Description",
    "Who", "What", "When", "Where", "Why",
    "Evidence Criteria", "Risk Rating", "System", "Role",
    "Regulatory Framework", "Status",
]
ws1.append(headers)
rows1 = [
    ["CTL-001","Access Review","Retail Banking","1.0 Lending",
     "Preventive","Automated","Daily","Daily access review for lending system",
     "IT Security Team","Verify user permissions","Daily at EOD","LOS","Prevent unauthorized access",
     "Access logs","High","Loan Origination System","IT Security Analyst","SOX","Active"],
    ["CTL-002","Reconciliation","Risk Management","2.0 Treasury",
     "Detective","Manual","Monthly","Monthly treasury reconciliation",
     "Treasury Ops","Reconcile balances","Month-end","Bloomberg","Ensure accuracy",
     "Recon reports","Medium","Bloomberg Terminal","Treasury Analyst","Basel III","Active"],
    ["CTL-003","Segregation of Duties","Compliance","3.0 Compliance",
     "Preventive","Manual","Quarterly","Quarterly SOD review",
     "Compliance Officer","Review SOD matrix","Quarterly","GRC Tool","Prevent conflicts",
     "SOD matrix","High","RSA Archer","Compliance Analyst","OCC","Active"],
    ["CTL-004","Change Management","IT Operations","4.0 Technology",
     "Preventive","Automated with Manual Component","Weekly","Weekly change review",
     "Change Manager","Approve changes","Every Friday","ServiceNow","Change control",
     "Change tickets","Low","ServiceNow","Change Manager","ITIL","Active"],
    ["CTL-005","Data Backup","IT Operations","4.0 Technology",
     "Contingency Planning","Automated","Daily","Daily data backup",
     "DBA","Backup databases","Nightly","AWS S3","Disaster recovery",
     "Backup logs","Medium","AWS S3","Database Administrator","ISO 27001","Active"],
]
for r in rows1:
    ws1.append(r)

ws2 = wb.create_sheet("section_2")
ws2.append(headers)
rows2 = [
    ["CTL-006","Transaction Monitoring","AML/BSA","5.0 Compliance",
     "Detective","Automated","Daily","Daily transaction monitoring",
     "AML Analyst","Monitor transactions","Real-time","Actimize","AML compliance",
     "Alert reports","High","Actimize","AML Analyst","BSA","Active"],
    ["CTL-007","Vendor Risk Assessment","Third Party Risk","6.0 Vendor",
     "Preventive","Manual","Annual","Annual vendor risk assessment",
     "TPRM Manager","Assess vendor risk","Annually","VRM Tool","Vendor risk",
     "Assessment reports","Medium","Archer","TPRM Analyst","OCC Guidance","Active"],
    ["CTL-008","Interest Rate Validation","Finance","7.0 Finance",
     "Detective","Automated","Daily","Rate validation check",
     "Finance Manager","Validate rates","Daily","Core Banking","Rate accuracy",
     "Validation logs","Low","Fiserv","Financial Analyst","GAAP","Active"],
    ["CTL-009","Loan Approval","Credit Risk","1.0 Lending",
     "Preventive","Manual","Per Transaction","Loan approval workflow",
     "Credit Officer","Approve loans","Per application","LOS","Credit risk",
     "Approval docs","High","Loan Origination System","Senior Credit Officer","OCC","Active"],
    ["CTL-010","Fraud Detection","Fraud Prevention","8.0 Fraud",
     "Detective","Automated","Daily","Real-time fraud scoring",
     "Fraud Analyst","Score transactions","Real-time","FICO","Fraud prevention",
     "Alert reports","High","FICO Falcon","Fraud Analyst","Dodd-Frank","Active"],
]
for r in rows2:
    ws2.append(r)

wb.save(FIXTURE_DIR / "sample_register.xlsx")

# ── nonstandard_register.xlsx: variant headers, 1 sheet ──
wb2 = openpyxl.Workbook()
ws = wb2.active
ws.title = "Sheet1"
headers2 = [
    "Ctrl ID", "Type of Control", "Division", "Process Area",
    "Preventive or Detective", "Automated/Manual", "How Often", "Narrative",
    "Responsible Party", "Action", "Timing", "Application", "Objective",
    "Evidence Required", "Risk Level",
]
ws.append(headers2)
rows3 = [
    ["C-1","SOD Review","Operations","Lending",
     "Preventive","Manual","Monthly","Monthly SOD review for ops",
     "Ops Manager","Review duty assignments","Monthly","Archer","Prevent conflicts",
     "SOD report","Medium"],
    ["C-2","Data Quality","Data Governance","Analytics",
     "Detective","Automated","Weekly","Weekly data quality scan",
     "Data Steward","Scan data quality","Weekly","Informatica","Data accuracy",
     "Quality dashboard","Low"],
    ["C-3","Patch Management","IT Security","Infrastructure",
     "Preventive","Automated","Monthly","Monthly patching cycle",
     "IT Admin","Apply patches","Monthly","WSUS","Security hygiene",
     "Patch report","High"],
]
for r in rows3:
    ws.append(r)

wb2.save(FIXTURE_DIR / "nonstandard_register.xlsx")
print("Done: created 2 fixture Excel files")
