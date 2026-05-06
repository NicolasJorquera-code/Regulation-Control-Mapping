"""Tests for Risk Inventory Builder."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from pydantic import ValidationError

from controlnexus.risk_inventory.calculators import (
    ControlEnvironmentCalculator,
    InherentRiskCalculator,
    ResidualRiskCalculator,
)
from controlnexus.risk_inventory.config import MatrixConfigLoader
from controlnexus.risk_inventory.demo import (
    financial_institution_demo_fixture_path,
    load_demo_risk_inventory,
    load_demo_workspace,
    load_knowledge_pack,
)
from controlnexus.risk_inventory.document_ingest import analyze_process_document
from controlnexus.risk_inventory.export import (
    build_hitl_review_workbook,
    build_risk_inventory_workspace_workbook,
    export_risk_inventory_to_excel,
    risk_inventory_excel_bytes,
    risk_inventory_review_excel_bytes,
    risk_inventory_workspace_excel_bytes,
)
from controlnexus.risk_inventory.graph import build_risk_inventory_graph
from controlnexus.risk_inventory.models import (
    ControlEffectivenessRating,
    ControlEnvironmentAssessment,
    ControlEnvironmentRating,
    ImpactAssessment,
    ImpactDimension,
    ImpactDimensionAssessment,
    ImpactScore,
    InherentRiskAssessment,
    Process,
    ReviewDecision,
    ReviewStatus,
    ApprovalStatus,
    RiskRating,
    RiskInventoryWorkspace,
)
from controlnexus.risk_inventory.services import (
    apply_review_decisions,
    build_control_gaps,
    build_synthetic_control_recommendations,
    run_risk_inventory_workflow,
    validate_knowledge_pack,
)
from controlnexus.risk_inventory.taxonomy import load_risk_inventory_taxonomy, load_root_cause_taxonomy
from controlnexus.risk_inventory.tools import build_risk_inventory_tool_executor
from controlnexus.risk_inventory.validator import RiskInventoryValidator


class TestRiskInventoryConfig:
    def test_config_loads(self):
        loader = MatrixConfigLoader()
        assert loader.impact_scales()["dimensions"]
        assert loader.frequency_scale()["scale"]
        assert loader.likelihood_scale()["scale"]
        assert loader.inherent_matrix()["matrix"]
        assert loader.residual_matrix()["matrix"]

    def test_taxonomy_crosswalk_has_required_categories(self):
        nodes = load_risk_inventory_taxonomy()
        categories = {node.level_2_category for node in nodes}
        required = {
            "Business Process Risk",
            "Data Management Risk",
            "IT Security / Cybersecurity Risk",
            "Regulatory Reporting Risk",
            "Third Party Risk",
        }
        assert required.issubset(categories)

    def test_root_cause_taxonomy_loads_basel_aligned_catalog(self):
        roots = load_root_cause_taxonomy()
        names = {root.name for root in roots}

        assert len(roots) == 22
        assert {root.category for root in roots} == {"External", "People", "Process", "Technology"}
        assert "Third-Party Service Failure" in names
        assert "Design Failure (Process, Control, or Policy)" in names
        assert "Inadequate Change, Release, or Deployment Practices" in names
        assert all(root.selection_criteria for root in roots)


class TestRiskInventoryCalculators:
    @pytest.mark.parametrize(
        ("impact", "likelihood", "expected"),
        [
            (1, 1, "Low-1"),
            (1, 4, "Medium-4"),
            (2, 3, "Medium-6"),
            (2, 4, "High-8"),
            (3, 3, "High-9"),
            (3, 4, "Critical-12"),
            (4, 1, "Medium-4"),
            (4, 4, "Critical-16"),
        ],
    )
    def test_inherent_matrix_key_cases(self, impact, likelihood, expected):
        result = InherentRiskCalculator().calculate(impact, likelihood)
        assert result.inherent_label == expected

    def test_inherent_matrix_all_16_combinations(self):
        calc = InherentRiskCalculator()
        expected = {
            1: {1: "Low-1", 2: "Low-2", 3: "Low-3", 4: "Medium-4"},
            2: {1: "Low-2", 2: "Medium-4", 3: "Medium-6", 4: "High-8"},
            3: {1: "Low-3", 2: "Medium-6", 3: "High-9", 4: "Critical-12"},
            4: {1: "Medium-4", 2: "High-8", 3: "Critical-12", 4: "Critical-16"},
        }
        for impact, likelihoods in expected.items():
            for likelihood, label in likelihoods.items():
                result = calc.calculate(impact, likelihood)
                assert result.inherent_score == impact * likelihood
                assert result.inherent_label == label

    @pytest.mark.parametrize(
        ("inherent_label", "environment", "expected"),
        [
            ("Low-1", "Strong", "Low-1"),
            ("Medium-4", "Inadequate", "Medium-20"),
            ("High-8", "Improvement Needed", "Medium-24"),
            ("Critical-12", "Satisfactory", "Medium-24"),
            ("Critical-16", "Inadequate", "Critical-80"),
        ],
    )
    def test_residual_matrix_key_cases(self, inherent_label, environment, expected):
        rating, score = inherent_label.split("-")
        inherent = InherentRiskAssessment(
            impact_score=ImpactScore.MINIMAL,
            likelihood_score=1,
            inherent_score=int(score),
            inherent_rating=RiskRating(rating),
            inherent_label=inherent_label,
        )
        env = ControlEnvironmentAssessment(
            design_rating=ControlEffectivenessRating(environment),
            operating_rating=ControlEffectivenessRating(environment),
            control_environment_rating=ControlEnvironmentRating(environment),
            rationale="test",
        )
        result = ResidualRiskCalculator().calculate(inherent, env)
        assert result.residual_label == expected

    def test_residual_matrix_all_configured_combinations(self):
        expected = {
            "Low-1": {"Strong": "Low-1", "Satisfactory": "Low-2", "Improvement Needed": "Low-3", "Inadequate": "Low-5"},
            "Low-2": {"Strong": "Low-2", "Satisfactory": "Low-4", "Improvement Needed": "Low-6", "Inadequate": "Low-10"},
            "Low-3": {"Strong": "Low-3", "Satisfactory": "Low-6", "Improvement Needed": "Low-9", "Inadequate": "Low-15"},
            "Medium-4": {"Strong": "Low-4", "Satisfactory": "Low-8", "Improvement Needed": "Low-12", "Inadequate": "Medium-20"},
            "Medium-6": {"Strong": "Low-6", "Satisfactory": "Low-12", "Improvement Needed": "Low-18", "Inadequate": "Medium-30"},
            "High-8": {"Strong": "Low-8", "Satisfactory": "Low-16", "Improvement Needed": "Medium-24", "Inadequate": "High-40"},
            "High-9": {"Strong": "Low-9", "Satisfactory": "Low-18", "Improvement Needed": "Medium-27", "Inadequate": "High-45"},
            "Critical-12": {"Strong": "Low-12", "Satisfactory": "Medium-24", "Improvement Needed": "High-36", "Inadequate": "Critical-60"},
            "Critical-16": {"Strong": "Low-16", "Satisfactory": "Medium-32", "Improvement Needed": "High-48", "Inadequate": "Critical-80"},
        }
        calc = ResidualRiskCalculator()
        for inherent_label, environments in expected.items():
            rating, score = inherent_label.split("-")
            inherent = InherentRiskAssessment(
                impact_score=ImpactScore.MINIMAL,
                likelihood_score=1,
                inherent_score=int(score),
                inherent_rating=RiskRating(rating),
                inherent_label=inherent_label,
            )
            for environment, label in environments.items():
                env = ControlEnvironmentAssessment(
                    design_rating=ControlEffectivenessRating(environment),
                    operating_rating=ControlEffectivenessRating(environment),
                    control_environment_rating=ControlEnvironmentRating(environment),
                    rationale="test",
                )
                assert calc.calculate(inherent, env).residual_label == label

    def test_control_environment_uses_worse_rating(self):
        result = ControlEnvironmentCalculator().calculate(
            ControlEffectivenessRating.STRONG,
            ControlEffectivenessRating.IMPROVEMENT_NEEDED,
        )
        assert result.control_environment_rating.value == "Improvement Needed"


class TestRiskInventoryModels:
    def test_impact_overall_below_dimension_requires_justification(self):
        with pytest.raises(ValidationError):
            ImpactAssessment(
                dimensions=[
                    ImpactDimensionAssessment(
                        dimension=ImpactDimension.FINANCIAL,
                        score=ImpactScore.SEVERE,
                        rationale="Severe financial impact.",
                    )
                ],
                overall_impact_score=ImpactScore.MEANINGFUL,
                overall_impact_rationale="Lower overall score.",
            )

    def test_process_alias_accepts_procedure_payload(self):
        process = Process.model_validate(
            {
                "procedure_id": "PROC-LEGACY",
                "procedure_name": "Legacy Named Process",
                "bu_id": "BU-TEST",
                "apqc_crosswalk": {
                    "framework": "APQC Banking PCF",
                    "version": "7.2.2",
                    "process_id": "banking-test",
                    "process_name": "Normalize test process",
                    "confidence": 0.75,
                    "rationale": "Optional crosswalk only.",
                },
            }
        )
        workspace = RiskInventoryWorkspace.model_validate(
            {"workspace_id": "WS-TEST", "procedures": [process.model_dump()]}
        )

        assert process.process_id == "PROC-LEGACY"
        assert process.procedure_name == "Legacy Named Process"
        assert process.apqc_crosswalk["framework"] == "APQC Banking PCF"
        assert workspace.procedures[0].process_name == "Legacy Named Process"


class TestRiskInventoryDemoAndGraph:
    def test_document_ingest_extracts_process_context(self):
        sample_path = Path("sample_data/risk_inventory_demo/payment_exception_policy.md")
        analysis = analyze_process_document(sample_path.name, sample_path.read_bytes())
        assert analysis.process_name == "Payment Exception Handling"
        assert analysis.product == "High-value payment processing"
        assert "Business Process Risk" in analysis.detected_risk_categories
        assert analysis.detected_controls
        assert analysis.exposure_cues

    def test_demo_loader_returns_complete_run(self):
        run = load_demo_risk_inventory()
        assert run.demo_mode is True
        assert run.input_context.process_name == "Payment Exception Handling"
        assert len(run.records) >= 6
        assert all(record.control_mappings for record in run.records)
        assert all(record.residual_risk.residual_label for record in run.records)

    def test_demo_loader_uses_specific_fixture_evidence(self):
        run = load_demo_risk_inventory()
        release = next(record for record in run.records if record.risk_id == "RISK-PAYEXC-001")

        assert "public payment case materials" in release.applicability.rationale
        assert release.residual_risk.management_response.recommended_action.startswith("Embed mandatory pre-release")
        assert any(metric.metric_name == "High-value repaired wires reviewed" for metric in release.exposure_metrics)
        assert any(reference.evidence_id == "EV-PAYEXC-001" for reference in release.evidence_references)
        assert run.run_manifest["evidence_metric_count"] >= 20
        assert run.run_manifest["scenario_basis"]
        assert "high-value erroneous wire-release scenario" in run.executive_summary.headline
        assert "Root-cause lens" not in release.risk_statement.risk_description

    def test_demo_requires_no_llm_and_validates(self):
        run = load_demo_risk_inventory()
        assert run.run_manifest["llm_required"] is False
        findings = RiskInventoryValidator().validate_run(run)
        assert all(finding.severity.value != "error" for finding in findings)

    def test_payment_demo_fixture_has_realistic_public_source_trace(self):
        body = Path("sample_data/risk_inventory_demo/payment_exception_handling.yaml").read_text(encoding="utf-8")
        run = load_demo_risk_inventory()

        assert "Generic demo evidence profile calibrated for executive walkthrough." not in body
        assert "Generic source pack" not in body
        assert "fictional" not in body.lower()
        assert "fictitious" not in body.lower()
        assert "Second Circuit Citibank/Revlon" in body
        assert any("occ.gov" in source for source in run.input_context.source_documents)
        assert any("law.justia.com" in source for source in run.input_context.source_documents)

    def test_payment_demo_risk_statements_are_executive_grade(self):
        run = load_demo_risk_inventory()

        for record in run.records:
            statement = record.risk_statement.risk_description
            assert len(statement.split()) >= 55
            assert statement.count(".") >= 3
            assert "Root-cause lens" not in statement
            assert "may be released" not in statement[:45] or len(statement.split()) > 70

    def test_demo_workspace_is_single_payment_exception_process(self):
        workspace = load_demo_workspace()
        assert workspace.bank_name == "Enterprise Payment Operations"
        assert len(workspace.business_units) == 1
        assert len(workspace.processes) == 1
        assert len(workspace.runs) == 1
        assert len(workspace.bank_controls) == 9
        assert len(workspace.control_inventory) == 9
        assert sum(len(run.records) for run in workspace.runs) == 8
        assert workspace.processes[0].process_id == "PROC-PAY-EXCEPTION"
        assert workspace.knowledge_pack_manifest["auto_generate_missing_runs"] is False
        assert not validate_knowledge_pack(workspace)
        assert all(workspace.run_for_process(process.process_id) for process in workspace.processes)
        assert all(run.records for run in workspace.runs)
        assert workspace.processes[0].apqc_crosswalk
        assert workspace.kri_library
        assert all(kri.kri_id.startswith("KRI-PAYEXC-") for kri in workspace.kri_library)
        assert all("generic" not in kri.metric_definition.lower() for kri in workspace.kri_library)
        assert all(control.who for control in workspace.control_inventory)
        assert all(control.what for control in workspace.control_inventory)
        assert all(control.when for control in workspace.control_inventory)
        assert all(control.where for control in workspace.control_inventory)
        assert all(control.why for control in workspace.control_inventory)
        assert all(control.evidence for control in workspace.control_inventory)

    def test_financial_institution_workspace_fixture_loads(self):
        workspace = load_demo_workspace(financial_institution_demo_fixture_path())

        assert workspace.workspace_id == "WS-FI-OPERATING-MODEL-V2"
        assert len(workspace.business_units) == 13
        assert len(workspace.processes) == 60
        assert len(workspace.runs) == 60
        assert len(workspace.control_inventory) > 200
        assert workspace.run_for_process("PR-16").run_id == "DEMO-FI-PR16-PAYEXC-001"
        assert workspace.run_for_process("PR-16").input_context.process_name == (
            "Payment Processing and Settlement (Wire/ACH)"
        )
        assert not validate_knowledge_pack(workspace)

    def test_knowledge_pack_loader_alias(self):
        workspace = load_knowledge_pack()
        assert workspace.workspace_id == "WS-PAYMENT-EXCEPTION-2026Q2"
        assert workspace.knowledge_pack_manifest["process_count"] == 1

    def test_loader_accepts_fixture_path_entries(self):
        workspace = load_knowledge_pack("sample_data/risk_inventory_demo/workspace.yaml")
        assert workspace.runs
        assert {run.input_context.process_id for run in workspace.runs} == {
            process.process_id for process in workspace.processes
        }

    def test_demo_loader_source_documents_use_fixture_name(self):
        run = load_demo_risk_inventory("sample_data/risk_inventory_demo/customer_onboarding.yaml")
        assert run.input_context.source_documents == ["risk inventory fixture: customer_onboarding.yaml"]

    def test_graph_deterministic_path(self):
        graph = build_risk_inventory_graph().compile()
        result = graph.invoke(
            {
                "run_id": "TEST-RI-001",
                "tenant_id": "test",
                "process_context": {
                    "process_id": "PROC-PAY-EXCEPTION",
                    "process_name": "Payment Exception Handling",
                    "product": "High-value payment processing",
                    "business_unit": "Payment Operations",
                    "description": "Daily high-value payment exception process with queue, reconciliation, access, and reporting.",
                    "systems": ["Payment Exception Workflow"],
                    "stakeholders": ["Operations Manager"],
                },
                "control_inventory": [
                    {
                        "control_id": "CTRL-PAY-001",
                        "control_name": "Daily exception queue review",
                        "control_type": "Exception Reporting",
                        "description": "Daily queue review and escalation.",
                    }
                ],
                "max_risks": 3,
            },
            config={"recursion_limit": 200},
        )
        report = result["final_report"]
        assert report["run_id"] == "TEST-RI-001"
        assert len(report["records"]) == 3
        assert report["records"][0]["inherent_risk"]["inherent_label"]

    def test_flagship_workflow_returns_trace_with_no_llm(self):
        workspace = load_demo_workspace()
        run = run_risk_inventory_workflow(
            workspace,
            {"process_id": "PROC-PAY-EXCEPTION"},
            llm_enabled=False,
        )
        assert run.input_context.process_id == "PROC-PAY-EXCEPTION"
        assert run.events
        assert all(event["mode"] in {"deterministic_fallback", "skipped"} for event in run.events)

    def test_synthetic_control_recommendations_for_gap(self):
        workspace = load_demo_workspace()
        record = next(
            record
            for run in workspace.runs
            for record in run.records
            if build_control_gaps(record)
        )

        gaps = build_control_gaps(record)
        recommendations = build_synthetic_control_recommendations(record, workspace)

        assert gaps
        assert recommendations
        assert recommendations[0].risk_id == record.risk_id
        assert recommendations[0].control_statement

    def test_risk_inventory_tool_executor_lookup(self):
        workspace = load_demo_workspace()
        execute = build_risk_inventory_tool_executor(workspace)

        process_result = execute(
            "knowledge_base_lookup",
            {"entity_type": "process", "entity_id": "PROC-PAY-EXCEPTION"},
        )
        matrix_result = execute("scoring_matrix_lookup", {"matrix": "residual"})

        assert process_result["records"][0]["process_name"] == "Payment Exception Handling"
        assert "matrix" in matrix_result


class TestRiskInventoryControlMappingUiHelpers:
    def test_demo_tab_list_removes_consolidated_tabs(self):
        from controlnexus.ui import risk_inventory_tab

        assert risk_inventory_tab.DEMO_RISK_INVENTORY_TABS == [
            "Knowledge Base",
            "Risk Inventory",
            "Control Mapping",
            "Gap Analysis",
        ]
        assert risk_inventory_tab.USER_RISK_INVENTORY_TABS == [
            "Knowledge Base",
            "Risk Inventory",
            "Control Mapping",
            "Gap Analysis",
        ]
        assert "Overview" not in risk_inventory_tab.USER_RISK_INVENTORY_TABS
        assert "Input / Upload" not in risk_inventory_tab.USER_RISK_INVENTORY_TABS
        assert "Control Gap Lab" not in risk_inventory_tab.DEMO_RISK_INVENTORY_TABS
        assert "Process Map" not in risk_inventory_tab.DEMO_RISK_INVENTORY_TABS
        assert "Residual Risk" not in risk_inventory_tab.DEMO_RISK_INVENTORY_TABS
        assert "KRI Program" not in risk_inventory_tab.DEMO_RISK_INVENTORY_TABS
        assert "Agent Run Trace" not in risk_inventory_tab.DEMO_RISK_INVENTORY_TABS
        assert "Review & Challenge" not in risk_inventory_tab.DEMO_RISK_INVENTORY_TABS
        assert "Review & Challenge" not in risk_inventory_tab.USER_RISK_INVENTORY_TABS
        assert "Executive Report" not in risk_inventory_tab.DEMO_RISK_INVENTORY_TABS
        assert "Executive Report" not in risk_inventory_tab.USER_RISK_INVENTORY_TABS

    def test_knowledge_base_tabs_remove_readiness_evidence_and_issues(self):
        from controlnexus.ui import risk_inventory_tab

        assert "Knowledge Pack Readiness" not in risk_inventory_tab.KNOWLEDGE_BASE_TABS
        assert "Evidence" not in risk_inventory_tab.KNOWLEDGE_BASE_TABS
        assert "Issues" not in risk_inventory_tab.KNOWLEDGE_BASE_TABS
        assert "Obligations" not in risk_inventory_tab.KNOWLEDGE_BASE_TABS

    def test_user_knowledge_base_intro_uses_plain_capability_rows(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")
        intro_source = source.split("def _render_user_knowledge_base_intro", 1)[1].split(
            "def _render_input_and_maybe_run", 1
        )[0]

        assert "Production Knowledge Base" not in intro_source
        assert "End-to-end output" not in intro_source
        assert "ri-kb-hero" not in source
        assert "ri-kb-page-title" in intro_source
        assert "Start with source evidence. Finish with a reviewer-ready risk inventory." in intro_source
        assert "Review what the workflow can consume" in intro_source
        assert "ri-kb-capability-grid" in intro_source
        assert "ri-kb-capability-row" in intro_source
        assert "Input Data" in intro_source
        assert "What the Workflow Consumes" in intro_source
        assert "Deliverables" in intro_source
        assert "What the Workflow Produces" in intro_source
        assert "Why It Helps" not in intro_source
        assert "Operating Context" in intro_source
        assert "actual operating model" in intro_source
        assert "Process Evidence" in intro_source
        assert "failures, handoffs, dependencies, and review points" in intro_source
        assert "Control Baseline" in intro_source
        assert "genuine control gaps" in intro_source
        assert "Risk Framework" in intro_source
        assert "current governance model" in intro_source
        assert "Risk Inventory" in intro_source
        assert "Scoring Record" in intro_source
        assert "Control Coverage" in intro_source
        assert "Reviewer Package" in intro_source
        assert "Excel-ready" in intro_source
        assert "export fields" in intro_source
        assert "Executive-Ready Output" not in intro_source
        assert "ri-kb-io-number" not in source
        assert "ri-kb-io-tags" not in source
        assert "ri-kb-io-card" not in source
        assert "ri-kb-benefit-band" not in source
        assert "ri-kb-flow-step" not in source

    def test_front_facing_demo_ui_labels_use_title_case(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        required_labels = [
            "Top Residual Risks",
            "Recommended Actions",
            "Root Causes",
            "Affected Stakeholders",
            "Current Inherent Basis",
            "CRO Rationale.",
            "Escalation Path.",
            "Evidence References",
            "Optimal Full-Coverage Control Statement",
            "Risk Coverage Rationale",
            "Expected Evidence.",
            "Why This Control Exists",
            "Fields Requiring Review",
            "Detected Risk Categories",
            "Extracted Text Preview",
        ]
        for label in required_labels:
            assert label in source

        lower_case_label_regressions = [
            "Top residual risks",
            "Recommended actions",
            "Root causes",
            "Affected stakeholders",
            "Current inherent basis",
            "**Management response**",
            "**Management response:**",
            "CRO rationale.",
            "Escalation path.",
            "Evidence references",
            "Optimal full-coverage control statement",
            "Risk coverage rationale",
            "Expected evidence.",
            "Why this control exists",
            "Fields requiring review",
            "Detected risk categories",
            "Extracted text preview",
        ]
        for label in lower_case_label_regressions:
            assert label not in source

    def test_knowledge_base_profile_options_are_single_process(self):
        from controlnexus.ui import risk_inventory_tab

        assert risk_inventory_tab.knowledge_base_profile_options() == ["Financial Institution Demo"]
        assert "MOCK_" + "INSTITUTION_PROFILES" not in vars(risk_inventory_tab)

    def test_profile_workspace_variants_filter_source_pack(self):
        large = load_knowledge_pack("sample_data/risk_inventory_demo/workspace.yaml")
        local = load_demo_workspace("sample_data/risk_inventory_demo/workspace_local_regional_bank.yaml")
        payments = load_demo_workspace("sample_data/risk_inventory_demo/workspace_digital_payments_institution.yaml")

        assert large.bank_name == "Large Global Bank"
        assert local.bank_name == "Local Regional Bank"
        assert payments.bank_name == "Digital Payments Institution"
        assert len(local.processes) < len(large.processes)
        assert len(payments.business_units) < len(large.business_units)
        assert all(local.run_for_process(process.process_id) for process in local.processes)
        assert all(payments.run_for_process(process.process_id) for process in payments.processes)

    def test_demo_ui_places_contextual_title_and_tabs_above_scope_selector(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        render_source = source.split("def render_risk_inventory_tab", 1)[1].split(
            "# ---------------------------------------------------------------------------\n# Demo workspace", 1
        )[0]
        title_source = source.split("def _render_risk_inventory_page_title", 1)[1].split(
            "# ---------------------------------------------------------------------------\n# Demo workspace", 1
        )[0]
        demo_source = source.split("def _render_demo_workspace", 1)[1].split(
            "# ---------------------------------------------------------------------------\n# User workflow", 1
        )[0]
        assert "Demo Workspace" in title_source
        assert "Financial Institution Risk Inventory Workbench" in title_source
        assert "Risk Inventory Workflow Builder" in title_source
        assert "ri-app-title-main" in title_source
        assert "<h1>Risk Inventory Builder</h1>" not in title_source
        assert "_render_risk_inventory_page_title(demo_enabled)" in render_source
        assert demo_source.index("tabs = st.tabs(DEMO_RISK_INVENTORY_TABS)") < demo_source.index(
            '_render_demo_scope_selector(workspace, "knowledge_base")'
        )
        assert "_render_demo_scope_selector(workspace)" not in demo_source
        assert '_render_demo_scope_selector(workspace, "knowledge_base")' in demo_source
        assert '_render_demo_scope_selector(workspace, "risk_inventory")' in demo_source
        assert '_render_demo_scope_selector(workspace, "control_mapping")' in demo_source
        assert '_render_demo_scope_selector(workspace, "gap_analysis")' in demo_source
        assert '"Business Unit"' in demo_source
        assert '"Process"' in demo_source
        assert "Workspace Dashboard (no process focus)" in demo_source
        assert "_render_workspace_aggregated_inventory(workspace, selected_bu_id)" in demo_source
        assert "selected_run = workspace.runs[0]" not in demo_source

    def test_risk_inventory_command_view_removes_management_response_strip(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")
        combined_source = source.split("def _render_risk_inventory_combined", 1)[1].split(
            "def _risk_statement_display", 1
        )[0]

        assert "_render_risk_command_review_summary" not in combined_source

    def test_demo_process_scope_label_leads_with_name_not_number(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace(financial_institution_demo_fixture_path())
        process = next(item for item in workspace.processes if item.process_id == "PR-16")

        label = risk_inventory_tab._process_scope_label(process, process.bu_id)

        assert label == f"{process.process_name} (Primary)"
        assert process.process_id not in label

    def test_bu_risk_capture_rows_show_differentiation(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_knowledge_pack("sample_data/risk_inventory_demo/workspace.yaml")
        rows = risk_inventory_tab.bu_risk_capture_rows(workspace)

        assert len(rows) == len(workspace.business_units)
        assert all(row["Key Source Packs"] for row in rows)
        assert all(row["Capture Rationale"] for row in rows)
        assert len({row["Dominant Captured Risk Types"] for row in rows}) > 1

    def test_bu_divergence_rows_show_level_2_drivers(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_knowledge_pack("sample_data/risk_inventory_demo/workspace.yaml")
        rows = risk_inventory_tab.bu_risk_divergence_rows(workspace)

        assert rows
        assert all(row["Level 2 Risk Category"] for row in rows)
        assert all(row["Enterprise Category"] for row in rows)
        assert all(row["Representative Process"] for row in rows)
        assert "Dominant Risk Category" not in rows[0]

    def test_no_banned_demo_names_in_ui_constants_or_sample_data(self):
        from controlnexus.ui import risk_inventory_tab

        banned = (
            "North" + "star",
            "Meri" + "dian",
            "Harbor" + "line",
            "FICT" + "-FS",
            "north" + "star-demo",
            "demo" + "-bank",
        )
        ui_blob = repr(risk_inventory_tab.KNOWLEDGE_BASE_PROFILES)
        sample_blob = "\n".join(
            path.read_text(encoding="utf-8")
            for path in Path("sample_data/risk_inventory_demo").rglob("*")
            if path.is_file() and path.suffix in {".yaml", ".yml", ".md"}
        )

        assert not any(term in ui_blob for term in banned)
        assert not any(term in sample_blob for term in banned)

    def test_risk_workbench_rows_include_every_selected_run_risk(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None

        rows = risk_inventory_tab.risk_inventory_workbench_rows(run, workspace)

        assert len(rows) == len(run.records)
        assert {row["Risk Record ID"] for row in rows} == {record.risk_id for record in run.records}
        assert list(rows[0])[:8] == [
            "Risk Record ID",
            "Business Unit",
            "Process",
            "Risk Subcategory",
            "Risk Statement",
            "Enterprise Risk Category",
            "Impact",
            "Frequency",
        ]
        assert list(rows[0])[8] == "Inherent Risk"
        assert all(row["Risk Statement"] for row in rows)
        assert all(len(row["Risk Statement"].split()) >= 55 for row in rows)
        assert all("Frequency" in row for row in rows)
        assert all("Inherent Risk" in row for row in rows)
        assert all("Residual Risk" not in row for row in rows)
        assert all(isinstance(row["Gaps"], int) for row in rows)
        assert all(row["Validation Level"] for row in rows)
        assert all(row["Required Reviewer"] for row in rows)

    def test_selected_risk_detail_contains_consolidated_profile(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        record = run.records[0]

        detail = risk_inventory_tab.selected_risk_detail(record, workspace)

        expected_keys = {
            "risk_id",
            "business_unit",
            "risk_statement",
            "root_causes",
            "impact_score",
            "frequency_score",
            "inherent_risk",
            "residual_risk",
            "management_response",
            "mitigation_plan",
            "controls",
            "control_gaps",
            "synthetic_controls",
            "kris",
            "evidence",
            "issues",
            "review",
            "apqc_crosswalk",
            "validation",
        }
        assert expected_keys.issubset(detail)
        assert detail["risk_id"] == record.risk_id
        assert detail["business_unit"]
        assert detail["frequency_score"] == int(record.likelihood_assessment.likelihood_score)
        assert detail["mitigation_plan"] == record.residual_risk.management_response.recommended_action
        assert isinstance(detail["controls"], list)
        assert isinstance(detail["kris"], list)
        assert isinstance(detail["evidence"], list)
        assert isinstance(detail["review"], dict)
        assert detail["validation"]["validation_level"]
        assert "Design Failure (Process, Control, or Policy)" in detail["root_causes"]
        assert "Root cause selection reflects" in detail["risk_statement"]

    def test_root_cause_taxonomy_rows_use_canonical_catalog(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        rows = risk_inventory_tab.root_cause_taxonomy_rows(workspace)

        assert len(rows) == 22
        assert list(rows[0]) == [
            "Code",
            "Category",
            "Cause Origin",
            "Root Cause",
            "Definition",
            "Selection Criteria",
            "Examples",
        ]
        assert rows[0]["Root Cause"] == "Third-Party Service Failure"
        assert all(row["Selection Criteria"] for row in rows)

    def test_impact_frequency_heatmap_marks_selected_cell(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        record = run.records[0]

        rows = risk_inventory_tab.impact_frequency_heatmap_rows(record)
        selected = [row for row in rows if row["Selected"]]

        assert len(rows) == 16
        assert len(selected) == 1
        assert selected[0]["Impact"] == int(record.impact_assessment.overall_impact_score)
        assert selected[0]["Frequency"] == int(record.likelihood_assessment.likelihood_score)
        assert selected[0]["Label"] == record.inherent_risk.inherent_label

    def test_portfolio_heatmap_rows_aggregate_business_unit_and_category(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        rows = risk_inventory_tab.portfolio_heatmap_rows(workspace)
        payment_bu = next(bu for bu in workspace.business_units if bu.bu_name == "Payment Operations")
        payment_rows = risk_inventory_tab.portfolio_heatmap_rows(workspace, payment_bu.bu_id)
        category_count = len({record.taxonomy_node.level_1_category for run in workspace.runs for record in run.records})

        assert rows
        assert len(rows) == len(workspace.business_units) * category_count
        assert {row["Business Unit"] for row in payment_rows} == {"Payment Operations"}
        assert any(row["Risk Records"] > 0 for row in rows)
        assert all(row["Heat"] in {"None", "Low", "Elevated", "Medium", "High"} for row in rows)

    def test_workspace_aggregated_inventory_rows_are_table_ready(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        rows = risk_inventory_tab.workspace_aggregated_inventory_rows(workspace)

        assert len(rows) == sum(len(run.records) for run in workspace.runs)
        assert list(rows[0])[:5] == [
            "Risk Record ID",
            "Business Unit",
            "Process",
            "Enterprise Risk Category",
            "Risk Subcategory",
        ]
        assert all(row["Risk Subcategory"] for row in rows)
        assert all(row["Business Unit"] for row in rows)
        assert all(row["Process"] for row in rows)
        assert all(row["Enterprise Risk Category"] for row in rows)

    def test_synthetic_control_inventory_rows_use_generated_recommendations(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        rows = risk_inventory_tab.synthetic_control_inventory_rows(workspace)

        assert rows
        assert list(rows[0]) == [
            "Recommendation ID",
            "Risk Record ID",
            "Business Unit",
            "Process",
            "Risk Subcategory",
            "Control",
            "Control Type",
            "Priority",
            "Owner",
            "Frequency",
            "Control Statement",
            "Rationale",
            "Expected Evidence",
        ]
        assert all(row["Recommendation ID"].startswith("SYN-") for row in rows)
        assert all(row["Rationale"] for row in rows)
        assert all(row["Expected Evidence"] for row in rows)

    def test_control_taxonomy_section_rows_use_section_control_workbook_shape(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        rows = risk_inventory_tab.control_taxonomy_section_rows(workspace)

        assert len(rows) == 175
        assert list(rows[0]) == risk_inventory_tab.CONTROL_INVENTORY_DISPLAY_COLUMNS
        assert set(rows[0]) == {
            "Control ID",
            "Control Level 1",
            "Control Level 2",
            "Full Description",
        }
        assert "hierarchy_id" in risk_inventory_tab.CONTROL_INVENTORY_COLUMNS
        assert "leaf_name" in risk_inventory_tab.CONTROL_INVENTORY_COLUMNS
        assert "business_unit_id" in risk_inventory_tab.CONTROL_INVENTORY_COLUMNS
        assert "business_unit_name" in risk_inventory_tab.CONTROL_INVENTORY_COLUMNS
        assert "hierarchy_id" not in rows[0]
        assert "leaf_name" not in rows[0]
        assert "business_unit_id" not in rows[0]
        assert "business_unit_name" not in rows[0]
        assert all(row["Control ID"].startswith("CTRL-01") for row in rows)
        assert all(row["Control Level 1"] for row in rows)
        assert all(row["Control Level 2"] for row in rows)
        assert all(row["Full Description"] for row in rows)

    def test_control_register_dataset_rows_use_full_section_control_workbook_shape(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        rows = risk_inventory_tab.control_register_dataset_rows(workspace)

        assert len(rows) == 175
        assert list(rows[0]) == risk_inventory_tab.CONTROL_INVENTORY_COLUMNS
        assert {
            "hierarchy_id",
            "who",
            "what",
            "when",
            "where",
            "why",
            "quality_rating",
            "validator_passed",
            "evidence",
        }.issubset(rows[0])
        assert all(row["control_id"].startswith("CTRL-01") for row in rows)
        assert all(row["hierarchy_id"] for row in rows)
        assert all(row["who"] for row in rows)
        assert all(row["what"] for row in rows)
        assert all(row["when"] for row in rows)
        assert all(row["where"] for row in rows)
        assert all(row["why"] for row in rows)
        assert all(row["quality_rating"] for row in rows)
        assert all("validator_passed" in row for row in rows)
        assert all(row["evidence"] for row in rows)

    def test_control_register_tab_reuses_generated_controls_table_renderer(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")
        register_source = source.split("def _render_control_register_dataset", 1)[1].split(
            "# ---------------------------------------------------------------------------\n# Risk Inventory tab", 1
        )[0]

        assert "render_data_table(" in register_source
        assert "CONTROL_REGISTER_DEFAULT_COLUMNS" in register_source
        assert "all_columns=CONTROL_INVENTORY_COLUMNS" in register_source
        assert 'export_filename="controls_section_1_register.csv"' in register_source

    def test_workspace_risk_inventory_render_does_not_render_divergence_table(self, monkeypatch):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        calls = []

        monkeypatch.setattr(risk_inventory_tab.st, "markdown", lambda *args, **kwargs: None)
        monkeypatch.setattr(risk_inventory_tab, "_render_workspace_inventory_summary", lambda *args, **kwargs: None)
        monkeypatch.setattr(risk_inventory_tab, "_render_prominent_table", lambda rows: calls.append(("inventory", len(rows))))
        monkeypatch.setattr(risk_inventory_tab, "_render_portfolio_heatmap", lambda rows: calls.append(("heatmap", len(rows))))
        monkeypatch.setattr(
            risk_inventory_tab,
            "_render_bu_difference_cards",
            lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("BU breakdown should not render")),
        )
        monkeypatch.setattr(
            risk_inventory_tab,
            "bu_risk_divergence_rows",
            lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("divergence table should not render")),
        )

        risk_inventory_tab._render_workspace_aggregated_inventory(workspace, None)

        assert calls[0][0] == "inventory"
        assert all(call[0] != "breakdown" for call in calls)

    def test_process_risk_inventory_moves_residual_and_source_confidence_out(self):
        center_source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        assert '"Residual": row["Residual Risk"]' not in center_source
        assert "Source confidence" not in center_source
        assert "Decision Stack" not in center_source
        assert '"Inherent Risk": row["Inherent Risk"]' in center_source
        assert "Why this matters" not in center_source
        assert "st.columns([1.52, 0.82]" not in center_source
        assert "_render_inherent_risk_summary(record)" in center_source
        assert "_render_compact_kri_panel" not in center_source
        assert "compact_kri_rows" not in center_source
        assert "_render_selected_risk_kri_cards(record, workspace)" in center_source
        assert "Linked KRI Program" not in center_source

    def test_selected_risk_kri_rows_are_full_width_card_ready(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        record = run.records[0]

        rows = risk_inventory_tab.selected_risk_kri_rows(record, workspace)

        assert rows
        assert list(rows[0]) == [
            "KRI ID",
            "KRI",
            "Definition",
            "Owner",
            "Frequency",
            "Source",
            "Green",
            "Amber",
            "Red",
            "Rationale",
            "Escalation Path",
        ]
        assert all(row["KRI ID"] for row in rows)
        assert all(row["KRI"] for row in rows)
        assert all(row["Rationale"] for row in rows)
        assert all(row["Escalation Path"] for row in rows)

    def test_selected_risk_kri_cards_use_compact_grid_without_formula(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        assert 'st.columns(2, gap="medium")' in source
        assert "Linked KRI Program" not in source
        assert "ri-selected-kri-threshold-line" in source
        assert "ri-selected-kri-formula" not in source
        assert "ri-selected-kri-thresholds" not in source
        assert "ri-selected-kri-grid" not in source

    def test_risk_inventory_browser_shows_wrapped_risk_statement_column(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")
        browser_source = source.split("def _render_risk_inventory_browser", 1)[1].split(
            "def _render_risk_command_center", 1
        )[0]

        assert '"Risk Subcategory": row["Risk Subcategory"]' in browser_source
        assert '"Risk Statement": row["Risk Statement"]' in browser_source
        assert "row_height = max(156, _table_row_height(browser_rows))" in browser_source
        assert "height=min(760, 64 + len(browser_rows) * row_height)" in browser_source

    def test_selected_risk_scoring_rationale_is_concise_and_metric_backed(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        record = run.records[0]

        impact = risk_inventory_tab.scoring_rationale_text(record, "impact")
        frequency = risk_inventory_tab.scoring_rationale_text(record, "frequency")

        for text in (impact, frequency):
            sentences = risk_inventory_tab._split_sentences(text)
            assert 2 <= len(sentences) <= 3
            assert "Current-period indicator:" in text
        assert any(metric.metric_name in impact for metric in record.exposure_metrics)
        assert any(metric.metric_name in frequency for metric in record.exposure_metrics)

    def test_selected_risk_inherent_panel_contains_rationale_grid(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        assert "ri-inherent-rationale-grid" in source
        assert 'scoring_rationale_text(record, "impact")' in source
        assert 'scoring_rationale_text(record, "frequency")' in source
        assert "Inherent Risk Matrix" not in source
        assert 'st.markdown("**Impact x Frequency**")' not in source

    def test_gap_analysis_exposes_residual_calculation_components(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        record = run.records[0]

        row = risk_inventory_tab.residual_calculation_row(record)

        assert list(row) == ["Inherent Risk", "Control Score", "Residual Risk Score", "Rationale"]
        assert row["Inherent Risk"] == record.inherent_risk.inherent_rating.value
        assert row["Control Score"] == record.residual_risk.control_environment_score
        assert row["Residual Risk Score"] == risk_inventory_tab.risk_rating_scale_score(
            record.residual_risk.residual_rating.value
        )
        assert row["Rationale"]
        assert record.residual_risk.residual_label not in row["Rationale"]
        assert "Deterministic" not in row["Rationale"]

    def test_gap_analysis_ui_does_not_show_combined_residual_labels(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        assert "residual_label" not in source
        assert "Deterministic rationale" not in source
        assert "ri-residual-calc-strip" in source

    def test_risk_inventory_tab_does_not_include_redundant_detail_expanders(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        assert "Controls and Coverage" not in source
        assert "Evidence and Source Trace" not in source
        assert "Issues and Open Findings" not in source
        assert "Review and Challenge" not in source
        assert "_render_selected_risk_drawer" not in source

    def test_table_column_labels_are_professionalized(self):
        from controlnexus.ui import risk_inventory_tab

        assert risk_inventory_tab._display_column_label("gap_type") == "Gap Type"
        assert risk_inventory_tab._display_column_label("existing_control_ids") == "Existing Control IDs"
        assert risk_inventory_tab._display_column_label("kri_id") == "KRI ID"
        assert risk_inventory_tab._display_column_label("Risk Record ID") == "Risk Record ID"

    def test_existing_knowledge_base_business_units_omits_1lod_lead(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")
        existing_tables_source = source.split("def _render_user_existing_knowledge_tables", 1)[1].split(
            "def _render_source_pack_tiles", 1
        )[0]

        assert '"1LOD Lead"' not in existing_tables_source

    def test_process_header_facts_are_limited_to_business_unit_and_process(self, monkeypatch):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        captured = {}

        monkeypatch.setattr(risk_inventory_tab.st, "markdown", lambda *args, **kwargs: None)
        monkeypatch.setattr(risk_inventory_tab, "_render_fact_block", lambda facts: captured.update(facts))

        risk_inventory_tab._render_process_command_header(run, workspace)

        assert set(captured) == {"Business Unit", "Process"}
        assert captured["Business Unit"] == "Payment Operations"
        assert captured["Process"] == "Payment Exception Handling"

    def test_risk_header_focuses_on_statement_not_color_badges(self, monkeypatch):
        from controlnexus.ui import risk_inventory_tab

        record = load_demo_risk_inventory().records[0]
        rendered = []

        monkeypatch.setattr(
            risk_inventory_tab.st,
            "markdown",
            lambda body, **kwargs: rendered.append(body),
        )

        risk_inventory_tab._render_risk_header(record)

        body = rendered[0]
        assert "ri-risk-statement-focus" in body
        assert record.risk_statement.risk_description[:40] in body
        assert "Inherent Risk:" not in body
        assert "Residual Risk:" not in body
        assert "Management Response:" not in body

    def test_risk_inventory_browser_no_longer_renders_summary_strip(self, monkeypatch):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        rows = risk_inventory_tab.risk_inventory_workbench_rows(run, workspace)
        markdown_calls = []

        class Selection:
            rows: list[int] = []

        class Event:
            selection = Selection()

        monkeypatch.setattr(
            risk_inventory_tab.st,
            "markdown",
            lambda body, **kwargs: markdown_calls.append(body),
        )
        monkeypatch.setattr(risk_inventory_tab.st, "dataframe", lambda *args, **kwargs: Event())

        risk_inventory_tab._render_risk_inventory_browser(rows, run.records[0].risk_id)

        assert not any("ri-inventory-summary" in str(call) for call in markdown_calls)

    def test_review_dossier_helper_returns_validation_context(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        record = run.records[0]

        queue_rows = risk_inventory_tab.review_validation_rows(run, workspace)
        dossier = risk_inventory_tab.selected_review_dossier(record, workspace)
        validation = risk_inventory_tab.required_validation_level(record, workspace)

        assert len(queue_rows) == len(run.records)
        assert all(row["Validation Level"] for row in queue_rows)
        assert validation["required_reviewer"]
        assert dossier["validation"]["validation_level"]
        assert dossier["controls"] is not None
        assert dossier["gaps"] is not None
        assert dossier["source_trace"]
        assert dossier["checklist"]
        assert dossier["scoring_rationale"]
        assert "synthetic_controls" in dossier
        assert dossier["level_2_category"] == record.taxonomy_node.level_2_category
        assert "configured" in dossier["bank_context_alignment"].lower()

    def test_neutral_callout_helper_uses_grey_callout_class(self, monkeypatch):
        from controlnexus.ui import risk_inventory_tab

        calls = []

        monkeypatch.setattr(
            risk_inventory_tab.st,
            "markdown",
            lambda body, **kwargs: calls.append({"body": body, **kwargs}),
        )

        risk_inventory_tab._render_neutral_callout("Choose a process focus.")

        assert "ri-neutral-callout" in calls[0]["body"]
        assert calls[0]["unsafe_allow_html"] is True

    def test_workspace_control_mapping_rows_filter_by_business_unit(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_knowledge_pack("sample_data/risk_inventory_demo/workspace.yaml")
        all_rows = risk_inventory_tab._workspace_control_mapping_rows(workspace)
        payment_ops = next(bu for bu in workspace.business_units if bu.bu_name == "Payment Operations")
        payment_rows = risk_inventory_tab._workspace_control_mapping_rows(workspace, payment_ops.bu_id)

        assert len(all_rows) == sum(len(run.records) for run in workspace.runs)
        assert payment_rows
        assert len(payment_rows) < len(all_rows)
        assert {row["Business Unit"] for row in payment_rows} == {"Payment Operations"}
        assert all(row["Enterprise Risk Category"] for row in payment_rows)
        assert all(isinstance(row["Mapped Controls"], int) for row in payment_rows)

    def test_workspace_control_mapping_matrix_summarizes_bu_and_risk_category(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        rows = risk_inventory_tab._workspace_control_mapping_rows(workspace)
        matrix = risk_inventory_tab._workspace_control_mapping_matrix_rows(rows)
        category_rows = risk_inventory_tab._workspace_control_mapping_category_rows(rows)

        assert len(matrix) == len({row["Business Unit"] for row in rows})
        assert all(row["Risk Records"] > 0 for row in matrix)
        assert all("Mapped Controls" in row for row in matrix)
        assert any(row["Enterprise Risk Category"] == "Operational" for row in category_rows)
        assert sum(row["Risk Records"] for row in category_rows) == len(rows)

    def test_control_score_row_uses_existing_control_environment_score(self):
        from controlnexus.ui import risk_inventory_tab

        run = load_demo_risk_inventory()
        record = run.records[0]

        row = risk_inventory_tab.control_score_row(record)
        configured_scores = MatrixConfigLoader().residual_matrix()["control_environment_scores"]

        assert row["Control Score"] == configured_scores[record.control_environment.control_environment_rating.value]
        assert row["Control Score"] == record.residual_risk.control_environment_score
        assert row["Control Strength"] == record.control_environment.control_environment_rating.value
        assert row["Mapped Controls"] == len(record.control_mappings)
        assert row["Coverage Status"] == risk_inventory_tab._record_coverage_status(record)
        assert row["Rationale"] == record.control_environment.rationale

    def test_control_mapping_process_summary_strip_is_removed(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")
        control_mapping_source = source.split("def _render_control_mapping", 1)[1].split(
            "def control_statement_detail", 1
        )[0]

        assert "_render_control_mapping_run_summary" not in source
        assert "Process Control Mapping Summary" not in source
        assert "Risk-Control Links" not in control_mapping_source
        assert "Open Gaps" not in control_mapping_source

    def test_selected_control_mapping_renderer_restores_process_controls_table(self, monkeypatch):
        from controlnexus.ui import risk_inventory_tab

        run = load_demo_risk_inventory()
        rendered = []
        table_calls = []
        selectbox_calls = []

        monkeypatch.setattr(risk_inventory_tab, "_risk_selector", lambda *_: run.records[0])
        monkeypatch.setattr(risk_inventory_tab, "_render_risk_header", lambda *_: None)
        monkeypatch.setattr(risk_inventory_tab, "_render_table", lambda rows, *_args, **_kwargs: table_calls.append(rows))
        monkeypatch.setattr(
            risk_inventory_tab.st,
            "selectbox",
            lambda label, options, **kwargs: selectbox_calls.append((label, options, kwargs)) or options[0],
        )
        monkeypatch.setattr(
            risk_inventory_tab.st,
            "markdown",
            lambda body, **kwargs: rendered.append(str(body)),
        )
        monkeypatch.setattr(risk_inventory_tab.st, "caption", lambda *_args, **_kwargs: None)

        risk_inventory_tab._render_control_mapping(run)
        body = "\n".join(rendered)

        assert table_calls
        assert selectbox_calls
        assert selectbox_calls[0][0] == "Control"
        assert selectbox_calls[0][1] == [mapping.control_id for mapping in run.records[0].control_mappings]
        assert len(table_calls[0]) == sum(len(record.control_mappings) for record in run.records)
        assert "All Mapped Controls In This Process" in body
        assert body.rfind("All Mapped Controls In This Process") > body.rfind("Selected Risk Control Coverage")
        assert "ri-control-score-panel" in body
        assert "ri-score-" in body
        assert "Control Score" in body
        assert "Rationale" in body
        assert "ri-control-coverage-panel" in body
        assert body.count("ri-control-coverage-panel") == 1
        assert "ri-control-card" not in body
        assert "Optimal Full-Coverage Control Statement" in body
        assert "Evidence to Prove Full Coverage." in body
        assert "ri-control-assessment-grid" in body
        assert "payment" in body.lower()
        assert "validation" in body.lower()
        assert "Control statements translate mapped controls" not in body
        assert "Last Tested" not in body
        assert "Sample" not in body
        assert "Exceptions" not in body
        assert "Open Issues" not in body
        assert "Open issue context" not in body
        assert "Design and operating rationale" not in body
        assert "Generic demo evidence profile calibrated for executive walkthrough." not in body

    def test_run_control_mapping_rows_include_control_score_context(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.runs[0]
        rows = risk_inventory_tab._run_control_mapping_rows(run, workspace)

        assert rows
        assert all(row["Control Objective"] for row in rows)
        assert all(row["Risk Coverage Rationale"] for row in rows)
        assert all("Mapped Root Causes" in row for row in rows)
        assert all("Control Score" in row for row in rows)
        assert all("Control Strength" in row for row in rows)
        assert all("Coverage Status" in row for row in rows)
        assert all(row["selected_level_1"] for row in rows)
        assert all(row["selected_level_2"] for row in rows)
        assert all(row["who"] for row in rows)
        assert all(row["what"] for row in rows)
        assert all(row["when"] for row in rows)
        assert all(row["frequency"] for row in rows)
        assert all(row["where"] for row in rows)
        assert all(row["why"] for row in rows)
        assert all(row["quality_rating"] for row in rows)
        assert all("validator_passed" in row for row in rows)
        assert all(row["evidence"] for row in rows)
        assert all("Control Statement" not in row for row in rows)
        assert all("Expected Evidence" not in row for row in rows)
        assert all("Residual Risk Rating" not in row for row in rows)
        assert all("full_description" not in row for row in rows)
        assert all("business_unit_id" not in row for row in rows)
        assert all("business_unit_name" not in row for row in rows)
        assert all("validator_retries" not in row for row in rows)
        assert all("validator_failures" not in row for row in rows)
        assert all("Evidence Quality" not in row for row in rows)
        assert all("Last Tested" not in row for row in rows)
        assert all("Sample Size" not in row for row in rows)
        assert all("Exceptions Noted" not in row for row in rows)
        assert all("Evidence Notes" not in row for row in rows)
        assert all("Design Rationale" not in row for row in rows)
        assert all("Operating Rationale" not in row for row in rows)

    def test_control_mapping_statement_detail_uses_statement_format(self):
        from controlnexus.ui import risk_inventory_tab

        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None
        record = run.records[0]
        detail = risk_inventory_tab.control_statement_detail(
            record,
            record.control_mappings[0],
            workspace,
        )
        recommendation = build_synthetic_control_recommendations(record, workspace)[0]

        assert detail["owner"] == recommendation.suggested_owner
        assert detail["frequency"] == recommendation.frequency
        assert detail["control_type"] == recommendation.control_type
        assert detail["coverage_label"] == "Full coverage target"
        assert detail["control_statement"] == recommendation.control_statement
        assert "expected evidence" not in detail["control_statement"].lower()
        assert detail["expected_evidence"] == recommendation.expected_evidence

    def test_control_mapping_renders_process_wide_table(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        assert "All Mapped Controls In This Process" in source
        assert "_render_process_linked_controls_table(run, workspace)" in source
        assert "Control statements translate mapped controls" not in source

    def test_selected_risk_control_coverage_uses_neutral_panel_without_teal_left_rail(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        assert "ri-control-coverage-panel" in source
        assert "ri-control-card { border-left: 4px solid #009d9a; }" not in source
        assert '<div class="ri-control-card">' not in source

    def test_gap_analysis_selected_risk_gap_table_removed(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")

        assert "Selected Risk Gap Analysis" not in source
        assert "Improved Control Statement" in source

    def test_gap_analysis_metric_strip_removed(self):
        source = Path("src/controlnexus/ui/risk_inventory_tab.py").read_text(encoding="utf-8")
        gap_lab_source = source.split("def _render_control_gap_lab", 1)[1].split(
            "def _render_gap_analysis_export", 1
        )[0]

        assert '"Risks Reviewed"' not in gap_lab_source
        assert '"Control Gaps"' not in gap_lab_source
        assert '"Synthetic Controls"' not in gap_lab_source
        assert '"No-Control Risks"' not in gap_lab_source
        assert '"Scoped Risks"' not in gap_lab_source
        assert '"Suggested Controls"' not in gap_lab_source


class TestRiskInventoryExcelExport:
    def test_export_bytes_contains_required_sheets(self):
        data = risk_inventory_excel_bytes(load_demo_risk_inventory())
        assert len(data) > 1000

    def test_export_creates_expected_sheets(self, tmp_path):
        path = export_risk_inventory_to_excel(load_demo_risk_inventory(), tmp_path / "risk_inventory.xlsx")
        wb = openpyxl.load_workbook(path)
        assert {
            "Executive Summary",
            "BU Risk Heatmap",
            "Risk Inventory",
            "Risk Statements and Root Causes",
            "Control Coverage and Gaps",
            "Synthetic Control Recs",
            "Residual Risk Mgmt Actions",
            "KRI Program",
            "Source Trace Config Snapshot",
            "Inherent Risk Assessment",
            "Control Mapping",
            "Control Effectiveness",
            "Residual Risk Assessment",
            "Review and Challenge",
            "Scoring Matrices",
            "Configuration Snapshot",
            "Validation Findings",
        }.issubset(set(wb.sheetnames))
        assert wb["Risk Inventory"].data_validations.count > 0

    def test_workspace_export_creates_demo_artifact_sheets(self):
        workspace = load_demo_workspace()
        wb = build_risk_inventory_workspace_workbook(workspace)

        assert {
            "Cover",
            "Executive Summary",
            "BU Risk Heatmap",
            "BU Risk Breakdown",
            "Process Risk Inventory",
            "Risk Detail Dossier",
            "Control Gap Summary",
            "Synthetic Control Recs",
            "KRI Dashboard",
            "Review & Challenge",
            "Reviewer Decision Log",
            "Source Trace",
            "Config Snapshot",
        }.issubset(set(wb.sheetnames))
        process_ids = {
            row[1]
            for row in wb["Process Risk Inventory"].iter_rows(min_row=2, values_only=True)
            if row and row[1]
        }
        business_units = {
            row[0]
            for row in wb["Executive Summary"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        assert process_ids == {"PROC-PAY-EXCEPTION"}
        assert business_units == {"Payment Operations"}

    def test_workspace_export_has_formatting_validations_and_trace(self):
        workspace = load_demo_workspace()
        decision = ReviewDecision(
            risk_id=workspace.runs[0].records[0].risk_id,
            reviewer="Senior Risk Reviewer",
            review_status=ReviewStatus.APPROVED,
            approval_status=ApprovalStatus.APPROVED,
            challenge_comments="Approved after evidence review.",
            reviewer_rationale="Scoring is supported by configured evidence.",
            final_approved_value="Approve",
        )
        wb = build_risk_inventory_workspace_workbook(workspace, [decision])

        assert wb["Cover"]["A1"].value == "Risk Inventory Builder"
        assert wb["Cover"]["A1"].fill.fgColor.rgb in {"00161616", "FF161616"}
        assert wb["Review & Challenge"].data_validations.count > 0
        assert wb["Reviewer Decision Log"].data_validations.count > 0
        assert wb["KRI Dashboard"].max_row > 1
        assert wb["Source Trace"].max_row > 1
        assert wb["Config Snapshot"].max_row > 1
        trace_types = {
            row[0]
            for row in wb["Source Trace"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        assert "Scenario Basis" in trace_types
        heat_values = [
            cell.fill.fgColor.rgb
            for row in wb["BU Risk Heatmap"].iter_rows(min_row=2)
            for cell in row
            if cell.fill.fill_type == "solid"
        ]
        assert heat_values
        decisions = [
            row
            for row in wb["Reviewer Decision Log"].iter_rows(min_row=2, values_only=True)
            if row and row[0] == decision.risk_id
        ]
        assert decisions and decisions[0][3] == "Senior Risk Reviewer"

    def test_workspace_export_bytes_contains_artifact(self):
        data = risk_inventory_workspace_excel_bytes(load_demo_workspace())
        assert len(data) > 2000

    def test_hitl_review_workbook_has_review_assets(self):
        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None

        wb = build_hitl_review_workbook(run, workspace)

        assert {
            "HITL Cover",
            "Validation Queue",
            "Review Checklist",
            "Rationale Breakdown",
            "Control Suggestions",
            "Evidence KRI Trace",
            "Reviewer Decision Log",
        }.issubset(set(wb.sheetnames))
        assert wb["HITL Cover"]["A2"].value == "Human-in-the-Loop Review Workbook"
        assert wb["Validation Queue"].data_validations.count > 0
        assert wb["Review Checklist"].max_row > len(run.records)
        assert wb["Rationale Breakdown"].max_row == len(run.records) + 1
        assert wb["Reviewer Decision Log"].data_validations.count > 0

    def test_hitl_review_excel_bytes_contains_workbook(self):
        workspace = load_demo_workspace()
        run = workspace.run_for_process("PROC-PAY-EXCEPTION")
        assert run is not None

        data = risk_inventory_review_excel_bytes(run, workspace)
        wb = openpyxl.load_workbook(BytesIO(data))

        assert len(data) > 2000
        assert "Review Checklist" in wb.sheetnames
        assert "Final Decision" in [cell.value for cell in wb["Validation Queue"][1]]

    def test_review_decision_applied_to_export_run(self):
        run = load_demo_risk_inventory()
        decision = ReviewDecision(
            risk_id=run.records[0].risk_id,
            reviewer="Senior Risk Reviewer",
            review_status=ReviewStatus.APPROVED,
            approval_status=ApprovalStatus.APPROVED,
            challenge_comments="Approved after evidence review.",
            reviewer_rationale="Scoring is supported by the fixture evidence.",
            final_approved_value="Residual rating approved",
        )

        updated = apply_review_decisions(run, [decision])
        review = updated.records[0].review_challenges[0]

        assert review.reviewer == "Senior Risk Reviewer"
        assert review.review_status == ReviewStatus.APPROVED
        assert review.approval_status == ApprovalStatus.APPROVED
        assert review.final_approved_value == "Residual rating approved"

    def test_streamlit_download_keys_include_location(self, monkeypatch):
        from controlnexus.ui import risk_inventory_tab

        run = load_demo_risk_inventory()
        calls = []

        monkeypatch.setattr(risk_inventory_tab, "risk_inventory_excel_bytes", lambda received_run: b"xlsx")

        def fake_download_button(label, **kwargs):
            calls.append({"label": label, **kwargs})
            return False

        monkeypatch.setattr(risk_inventory_tab.st, "download_button", fake_download_button)

        risk_inventory_tab._download_export(run, "overview")
        risk_inventory_tab._download_export(run, "gap_analysis")

        keys = [call["key"] for call in calls]
        assert keys == [f"ri_xlsx_overview_{run.run_id}", f"ri_xlsx_gap_analysis_{run.run_id}"]
        assert len(set(keys)) == 2

    def test_review_tab_download_uses_hitl_workbook(self, monkeypatch):
        from controlnexus.ui import risk_inventory_tab

        run = load_demo_risk_inventory()
        calls = []

        monkeypatch.setattr(
            risk_inventory_tab,
            "risk_inventory_review_excel_bytes",
            lambda received_run, workspace, decisions: b"hitl-xlsx",
        )

        def fake_download_button(label, **kwargs):
            calls.append({"label": label, **kwargs})
            return False

        monkeypatch.setattr(risk_inventory_tab.st, "download_button", fake_download_button)
        monkeypatch.setattr(risk_inventory_tab.st, "caption", lambda *args, **kwargs: None)

        risk_inventory_tab._download_review_workbook(run)

        assert calls[0]["label"] == "Download HITL Review Workbook"
        assert calls[0]["key"] == f"ri_review_xlsx_{run.run_id}"
        assert calls[0]["file_name"] == f"{run.run_id}_hitl_review_workbook.xlsx"
